from fastapi import APIRouter, HTTPException, Body, Depends
from typing import Optional
from datetime import datetime, timezone
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging

from database import db
from models import Bobin, BobinMovement
from auth import get_current_user
from services.audit import log_audit

router = APIRouter(dependencies=[Depends(get_current_user)])
logger = logging.getLogger(__name__)


def bobin_label(b: dict) -> str:
    layers = int(b.get('layers', 1) or 1)
    layer_str = "TEK" if layers == 1 else "CIFT" if layers == 2 else f"{layers} KAT"
    return f"{b.get('brand', '')} {b.get('width_cm', '')}cm {b.get('grammage', '')}gr {b.get('color', '')} {layer_str}"


# ==================== BOBİN CRUD ====================

@router.get("/bobins")
async def get_bobins():
    bobins = await db.bobins.find({}, {"_id": 0}).sort("brand", 1).to_list(500)
    return bobins


@router.get("/bobins/barcode/{code}")
async def get_bobin_by_barcode(code: str):
    """Barkod ile bobin bul"""
    bobin = await db.bobins.find_one({"barcode": code}, {"_id": 0})
    if not bobin:
        raise HTTPException(status_code=404, detail="Bu barkoda ait bobin bulunamadi")
    return bobin


@router.post("/bobins")
async def create_bobin(data: dict = Body(...)):
    """Stoğa bobin ekle - kilogram bazlı (adet opsiyonel)."""
    brand = data.get("brand", "").strip()
    width_cm = float(data.get("width_cm", 0))
    grammage = float(data.get("grammage", 0))
    color = data.get("color", "Beyaz").strip()
    layers = int(data.get("layers", 1) or 1)
    if layers < 1:
        layers = 1
    quantity = int(data.get("quantity", 0) or 0)
    total_weight_kg = float(data.get("total_weight_kg", 0) or 0)
    supplier = data.get("supplier", "")
    notes = data.get("notes", "")
    user_name = data.get("user_name", "")
    barcode = data.get("barcode", "").strip()

    if not brand or width_cm <= 0 or grammage <= 0:
        raise HTTPException(status_code=400, detail="Marka, genislik ve gramaj zorunludur")
    if total_weight_kg <= 0:
        raise HTTPException(status_code=400, detail="Toplam agirlik (kg) sifirdan buyuk olmalidir")

    weight_per_piece = round(total_weight_kg / quantity, 2) if quantity > 0 else 0

    # Barkod ile arama (varsa)
    existing = None
    if barcode:
        existing = await db.bobins.find_one({"barcode": barcode}, {"_id": 0})

    # Barkod yoksa marka/olcu/gramaj/renk/kat ile ara
    if not existing:
        existing = await db.bobins.find_one({
            "brand": brand, "width_cm": width_cm,
            "grammage": grammage, "color": color,
            "layers": layers
        }, {"_id": 0})

    if existing:
        new_qty = (existing.get("quantity", 0) or 0) + quantity
        new_weight = (existing.get("total_weight_kg", 0) or 0) + total_weight_kg
        new_wpp = round(new_weight / new_qty, 2) if new_qty > 0 else 0
        update_data = {
            "quantity": new_qty,
            "total_weight_kg": round(new_weight, 2),
            "weight_per_piece_kg": new_wpp,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        if barcode and not existing.get("barcode"):
            update_data["barcode"] = barcode
        await db.bobins.update_one({"id": existing["id"]}, {"$set": update_data})
        label = bobin_label(existing)
        movement = BobinMovement(
            bobin_id=existing["id"], bobin_label=label,
            movement_type="purchase", quantity=quantity,
            weight_kg=total_weight_kg,
            note=f"Tedarikci: {supplier}" if supplier else notes,
            user_name=user_name
        )
        await db.bobin_movements.insert_one(movement.model_dump())
        await log_audit(user_name or "Depo", "purchase", "bobin", label,
                        f"+{total_weight_kg}kg, Tedarikci: {supplier}")
        updated = await db.bobins.find_one({"id": existing["id"]}, {"_id": 0})
        return {"bobin": updated, "message": f"{label} stoka eklendi (+{total_weight_kg:g} kg)"}
    else:
        bobin = Bobin(
            barcode=barcode, brand=brand, width_cm=width_cm, grammage=grammage, color=color,
            layers=layers,
            quantity=quantity, total_weight_kg=round(total_weight_kg, 2),
            weight_per_piece_kg=weight_per_piece, supplier=supplier, notes=notes
        )
        await db.bobins.insert_one(bobin.model_dump())
        label = bobin_label(bobin.model_dump())
        movement = BobinMovement(
            bobin_id=bobin.id, bobin_label=label,
            movement_type="purchase", quantity=quantity,
            weight_kg=total_weight_kg,
            note=f"Tedarikci: {supplier}" if supplier else notes,
            user_name=user_name
        )
        await db.bobin_movements.insert_one(movement.model_dump())
        await log_audit(user_name or "Depo", "create", "bobin", label,
                        f"{total_weight_kg}kg, Tedarikci: {supplier}")
        return {"bobin": bobin.model_dump(), "message": f"{label} olusturuldu ({total_weight_kg:g} kg)"}


@router.patch("/bobins/{bobin_id}")
async def update_bobin(bobin_id: str, data: dict = Body(...)):
    """Yanlış girilmiş bobin bilgilerini düzenle (yetkili)."""
    bobin = await db.bobins.find_one({"id": bobin_id}, {"_id": 0})
    if not bobin:
        raise HTTPException(status_code=404, detail="Bobin bulunamadi")

    user_name = data.get("user_name", "Depo")
    update_fields = {}
    allowed_str = ["brand", "color", "barcode", "supplier", "notes"]
    allowed_num = ["width_cm", "grammage", "total_weight_kg"]
    allowed_int = ["quantity", "layers"]

    for k in allowed_str:
        if k in data:
            v = (data.get(k) or "").strip() if isinstance(data.get(k), str) else data.get(k)
            update_fields[k] = v
    for k in allowed_num:
        if k in data and data.get(k) not in (None, ""):
            try:
                update_fields[k] = float(data.get(k))
            except Exception:
                pass
    for k in allowed_int:
        if k in data and data.get(k) not in (None, ""):
            try:
                update_fields[k] = int(data.get(k))
            except Exception:
                pass

    if not update_fields:
        raise HTTPException(status_code=400, detail="Guncellenecek alan yok")

    # Yeni alanlardan adet basi kg yeniden hesapla
    qty = update_fields.get("quantity", bobin.get("quantity", 0)) or 0
    weight = update_fields.get("total_weight_kg", bobin.get("total_weight_kg", 0)) or 0
    update_fields["weight_per_piece_kg"] = round(weight / qty, 2) if qty > 0 else 0
    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.bobins.update_one({"id": bobin_id}, {"$set": update_fields})
    updated = await db.bobins.find_one({"id": bobin_id}, {"_id": 0})

    label_old = bobin_label(bobin)
    label_new = bobin_label(updated)
    diff = ", ".join([f"{k}: {bobin.get(k, '-')} -> {update_fields[k]}"
                      for k in update_fields if k not in ("updated_at", "weight_per_piece_kg")])
    await log_audit(user_name, "update", "bobin", label_new, f"{label_old} duzenlendi: {diff}")

    return {"bobin": updated, "message": f"{label_new} guncellendi"}


@router.delete("/bobins/{bobin_id}")
async def delete_bobin(bobin_id: str, data: dict = Body(None)):
    bobin = await db.bobins.find_one({"id": bobin_id}, {"_id": 0})
    if not bobin:
        raise HTTPException(status_code=404, detail="Bobin bulunamadi")
    if (bobin.get("total_weight_kg", 0) or 0) > 0:
        raise HTTPException(status_code=400, detail="Stokta agirlik varken silinemez")
    await db.bobins.delete_one({"id": bobin_id})
    user_name = data.get("user_name", "Depo") if data else "Depo"
    await log_audit(user_name, "delete", "bobin", bobin_label(bobin))
    return {"message": "Bobin silindi"}


# ==================== STOK GİRİŞ (SATIN ALMA) ====================

@router.post("/bobins/{bobin_id}/purchase")
async def purchase_bobin(bobin_id: str, data: dict = Body(...)):
    """Mevcut bobine kg ekle (adet opsiyonel)."""
    quantity = int(data.get("quantity", 0) or 0)
    weight_kg = float(data.get("weight_kg", 0) or 0)
    supplier = data.get("supplier", "")
    user_name = data.get("user_name", "")
    note = data.get("note", "")

    if weight_kg <= 0:
        raise HTTPException(status_code=400, detail="Agirlik (kg) sifirdan buyuk olmalidir")

    bobin = await db.bobins.find_one({"id": bobin_id}, {"_id": 0})
    if not bobin:
        raise HTTPException(status_code=404, detail="Bobin bulunamadi")

    new_qty = (bobin.get("quantity", 0) or 0) + quantity
    new_weight = (bobin.get("total_weight_kg", 0) or 0) + weight_kg
    new_wpp = round(new_weight / new_qty, 2) if new_qty > 0 else 0

    await db.bobins.update_one(
        {"id": bobin_id},
        {"$set": {
            "quantity": new_qty, "total_weight_kg": round(new_weight, 2),
            "weight_per_piece_kg": new_wpp,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    label = bobin_label(bobin)
    movement = BobinMovement(
        bobin_id=bobin_id, bobin_label=label,
        movement_type="purchase", quantity=quantity, weight_kg=weight_kg,
        note=f"Tedarikci: {supplier}. {note}".strip() if supplier else note,
        user_name=user_name
    )
    await db.bobin_movements.insert_one(movement.model_dump())
    await log_audit(user_name or "Depo", "purchase", "bobin", label, f"+{weight_kg}kg")
    return {"message": f"{label} stok eklendi (+{weight_kg:g} kg)", "new_weight": round(new_weight, 2)}


# ==================== MAKİNEYE VER ====================

@router.post("/bobins/{bobin_id}/to-machine")
async def give_bobin_to_machine(bobin_id: str, data: dict = Body(...)):
    """Bobini makineye/hedef noktaya ver - kg bazlı (adet opsiyonel)."""
    quantity = int(data.get("quantity", 0) or 0)
    weight_kg_input = data.get("weight_kg")
    machine_id = data.get("machine_id", "")
    machine_name = data.get("machine_name", "")
    user_name = data.get("user_name", "")

    bobin = await db.bobins.find_one({"id": bobin_id}, {"_id": 0})
    if not bobin:
        raise HTTPException(status_code=404, detail="Bobin bulunamadi")

    # Ağırlık hesapla: weight_kg verilmişse onu kullan, yoksa adet*wpp
    if weight_kg_input not in (None, ""):
        weight_out = float(weight_kg_input)
    else:
        wpp = bobin.get("weight_per_piece_kg", 0) or 0
        weight_out = round(wpp * quantity, 2)

    if weight_out <= 0:
        raise HTTPException(status_code=400, detail="Agirlik (kg) sifirdan buyuk olmalidir")

    current_weight = bobin.get("total_weight_kg", 0) or 0
    if weight_out > current_weight + 0.001:
        raise HTTPException(status_code=400, detail=f"Yetersiz stok! Mevcut: {current_weight:g}kg")

    new_qty = max((bobin.get("quantity", 0) or 0) - quantity, 0)
    new_weight = max(round(current_weight - weight_out, 2), 0)
    new_wpp = round(new_weight / new_qty, 2) if new_qty > 0 else 0

    await db.bobins.update_one(
        {"id": bobin_id},
        {"$set": {
            "quantity": new_qty, "total_weight_kg": new_weight,
            "weight_per_piece_kg": new_wpp,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    label = bobin_label(bobin)
    movement = BobinMovement(
        bobin_id=bobin_id, bobin_label=label,
        movement_type="to_machine", quantity=quantity, weight_kg=weight_out,
        machine_id=machine_id, machine_name=machine_name, user_name=user_name
    )
    await db.bobin_movements.insert_one(movement.model_dump())
    await log_audit(user_name or "Depo", "to_machine", "bobin", label,
                    f"-{weight_out}kg -> {machine_name}")
    return {"message": f"{weight_out:g} kg {label} -> {machine_name}", "new_weight": new_weight}


# ==================== MÜŞTERİYE SATIŞ ====================

@router.post("/bobins/{bobin_id}/sale")
async def sell_bobin(bobin_id: str, data: dict = Body(...)):
    """Müşteriye sat - kg bazlı (adet opsiyonel)."""
    quantity = int(data.get("quantity", 0) or 0)
    weight_kg_input = data.get("weight_kg")
    customer_name = data.get("customer_name", "")
    user_name = data.get("user_name", "")
    note = data.get("note", "")

    if not customer_name:
        raise HTTPException(status_code=400, detail="Musteri adi zorunludur")

    bobin = await db.bobins.find_one({"id": bobin_id}, {"_id": 0})
    if not bobin:
        raise HTTPException(status_code=404, detail="Bobin bulunamadi")

    if weight_kg_input not in (None, ""):
        weight_out = float(weight_kg_input)
    else:
        wpp = bobin.get("weight_per_piece_kg", 0) or 0
        weight_out = round(wpp * quantity, 2)

    if weight_out <= 0:
        raise HTTPException(status_code=400, detail="Agirlik (kg) sifirdan buyuk olmalidir")

    current_weight = bobin.get("total_weight_kg", 0) or 0
    if weight_out > current_weight + 0.001:
        raise HTTPException(status_code=400, detail=f"Yetersiz stok! Mevcut: {current_weight:g}kg")

    new_qty = max((bobin.get("quantity", 0) or 0) - quantity, 0)
    new_weight = max(round(current_weight - weight_out, 2), 0)
    new_wpp = round(new_weight / new_qty, 2) if new_qty > 0 else 0

    await db.bobins.update_one(
        {"id": bobin_id},
        {"$set": {
            "quantity": new_qty, "total_weight_kg": new_weight,
            "weight_per_piece_kg": new_wpp,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    label = bobin_label(bobin)
    movement = BobinMovement(
        bobin_id=bobin_id, bobin_label=label,
        movement_type="sale", quantity=quantity, weight_kg=weight_out,
        customer_name=customer_name, note=note, user_name=user_name
    )
    await db.bobin_movements.insert_one(movement.model_dump())
    await log_audit(user_name or "Depo", "sale", "bobin", label,
                    f"-{weight_out}kg -> Musteri: {customer_name}")
    return {"message": f"{weight_out:g} kg {label} -> {customer_name}", "new_weight": new_weight}


# ==================== HAREKET GEÇMİŞİ ====================

@router.get("/bobins/movements")
async def get_bobin_movements(bobin_id: Optional[str] = None, movement_type: Optional[str] = None, limit: int = 200):
    query = {}
    if bobin_id:
        query["bobin_id"] = bobin_id
    if movement_type:
        query["movement_type"] = movement_type
    movements = await db.bobin_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return movements


# ==================== EXCEL EXPORT (kapsamlı + aylık arşiv) ====================

def _layer_label(b):
    lyr = int(b.get("layers", 1) or 1)
    return "TEK" if lyr == 1 else "CIFT" if lyr == 2 else f"{lyr} KAT"


def _bobin_key(m):
    """Hareket kaydından bobin grup anahtarı türetir (label kullanır)."""
    return m.get("bobin_label", "") or m.get("bobin_id", "")


@router.get("/bobins/archive/months")
async def list_archive_months():
    """Hareket geçmişinde mevcut olan (YYYY-MM) listesini döndürür."""
    pipeline = [
        {"$project": {"month": {"$substr": ["$created_at", 0, 7]}}},
        {"$group": {"_id": "$month"}},
        {"$sort": {"_id": -1}},
        {"$limit": 36},
    ]
    cursor = db.bobin_movements.aggregate(pipeline)
    months = [doc["_id"] async for doc in cursor if doc.get("_id")]
    return months


@router.get("/bobins/export")
async def export_bobins(month: Optional[str] = None, data: dict = Depends(get_current_user)):
    """
    Bobin Excel raporu.
    - month=YYYY-MM verilmezse: ANLIK durum (mevcut stok + son 2000 hareket)
    - month=YYYY-MM verilirse: O AYIN ARŞİV RAPORU (özet + tüm hareketler + makine dağılımı + müşteri satışları)
    """
    user_display = data.get("display_name", data.get("username", "Bilinmeyen"))

    # ---- Tarih aralığı belirle ----
    is_archive = bool(month)
    period_label = ""
    movement_query = {}
    if is_archive:
        try:
            year, mo = month.split("-")
            year_i, mo_i = int(year), int(mo)
            if mo_i < 1 or mo_i > 12:
                raise ValueError
        except Exception:
            raise HTTPException(status_code=400, detail="Gecersiz ay formati. YYYY-MM bekleniyor.")
        period_start = datetime(year_i, mo_i, 1, tzinfo=timezone.utc)
        if mo_i == 12:
            period_end = datetime(year_i + 1, 1, 1, tzinfo=timezone.utc)
        else:
            period_end = datetime(year_i, mo_i + 1, 1, tzinfo=timezone.utc)
        movement_query = {"created_at": {"$gte": period_start.isoformat(), "$lt": period_end.isoformat()}}
        TR_AY = ["", "Ocak", "Subat", "Mart", "Nisan", "Mayis", "Haziran",
                 "Temmuz", "Agustos", "Eylul", "Ekim", "Kasim", "Aralik"]
        period_label = f"{TR_AY[mo_i]} {year_i}"
    else:
        period_start = None

    bobins = await db.bobins.find({}, {"_id": 0}).sort("brand", 1).to_list(500)
    movements = await db.bobin_movements.find(movement_query, {"_id": 0}).sort("created_at", 1).to_list(20000)

    # ---- Stiller ----
    wb = Workbook()
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    sub_fill = PatternFill(start_color="FFF3D6", end_color="FFF3D6", fill_type="solid")
    bold = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, cols, fill=None):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = fill or header_fill
            cell.alignment = center_align

    def auto_width(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def title_row(ws, span, text):
        ws.merge_cells(f"A1:{get_column_letter(span)}1")
        c = ws.cell(row=1, column=1, value=text)
        c.font = title_font
        c.fill = title_fill
        c.alignment = center_align
        ws.row_dimensions[1].height = 30

    # ==================== SAYFA 1: ÖZET ====================
    ws1 = wb.active
    ws1.title = "Ozet"
    if is_archive:
        # Aylık özet: bobin başına ay başı stok / ay içi giriş / ay içi çıkış / ay sonu stok
        title_row(ws1, 9, f"BUSE KAGIT - Bobin Aylik Ozet ({period_label})")

        # Tarih öncesi tüm hareketler (ay başı stok hesaplama için)
        pre_movements = await db.bobin_movements.find(
            {"created_at": {"$lt": period_start.isoformat()}}, {"_id": 0}
        ).to_list(50000)

        # Bobin label başına ay başı / ay içi toplam
        opening = {}  # label -> kg
        for m in pre_movements:
            k = _bobin_key(m)
            if not k:
                continue
            sign = 1 if m.get("movement_type") == "purchase" else -1
            opening[k] = opening.get(k, 0) + sign * float(m.get("weight_kg", 0) or 0)

        in_month_in = {}   # purchase
        in_month_out_machine = {}
        in_month_out_sale = {}
        labels_seen = set(opening.keys())
        for m in movements:
            k = _bobin_key(m)
            if not k:
                continue
            labels_seen.add(k)
            w = float(m.get("weight_kg", 0) or 0)
            mt = m.get("movement_type")
            if mt == "purchase":
                in_month_in[k] = in_month_in.get(k, 0) + w
            elif mt == "to_machine":
                in_month_out_machine[k] = in_month_out_machine.get(k, 0) + w
            elif mt == "sale":
                in_month_out_sale[k] = in_month_out_sale.get(k, 0) + w

        headers = ["Bobin", "Ay Basi Stok (kg)", "Ay Ici Giris (kg)",
                   "Ay Ici Makineye (kg)", "Ay Ici Satis (kg)",
                   "Ay Sonu Stok (kg)", "Net Degisim (kg)", "Hareket Sayisi", "Aktif?"]
        for col, h in enumerate(headers, 1):
            ws1.cell(row=3, column=col, value=h)
        style_header(ws1, 3, len(headers))

        sorted_labels = sorted(labels_seen)
        active_set = {bobin_label(b) for b in bobins}
        row = 4
        sum_open = sum_in = sum_out_m = sum_out_s = sum_close = 0
        for k in sorted_labels:
            o = round(opening.get(k, 0), 2)
            inq = round(in_month_in.get(k, 0), 2)
            outm = round(in_month_out_machine.get(k, 0), 2)
            outs = round(in_month_out_sale.get(k, 0), 2)
            close = round(o + inq - outm - outs, 2)
            net = round(close - o, 2)
            hareket = sum(1 for m in movements if _bobin_key(m) == k)
            ws1.cell(row=row, column=1, value=k)
            ws1.cell(row=row, column=2, value=o).alignment = center_align
            ws1.cell(row=row, column=3, value=inq).alignment = center_align
            ws1.cell(row=row, column=4, value=outm).alignment = center_align
            ws1.cell(row=row, column=5, value=outs).alignment = center_align
            ws1.cell(row=row, column=6, value=close).alignment = center_align
            ws1.cell(row=row, column=7, value=net).alignment = center_align
            ws1.cell(row=row, column=8, value=hareket).alignment = center_align
            ws1.cell(row=row, column=9, value="Evet" if k in active_set else "Hayir").alignment = center_align
            sum_open += o
            sum_in += inq
            sum_out_m += outm
            sum_out_s += outs
            sum_close += close
            row += 1
        # Toplam satırı
        ws1.cell(row=row, column=1, value="TOPLAM").font = bold
        for c, v in [(2, sum_open), (3, sum_in), (4, sum_out_m), (5, sum_out_s), (6, sum_close), (7, sum_close - sum_open)]:
            cell = ws1.cell(row=row, column=c, value=round(v, 2))
            cell.alignment = center_align
            cell.font = bold
            cell.fill = sub_fill
        auto_width(ws1, [40, 18, 18, 22, 18, 18, 18, 14, 10])
    else:
        # Anlık stok
        title_row(ws1, 7, f"BUSE KAGIT - Bobin Stok Durumu ({datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')})")
        headers = ["Barkod", "Marka", "Genislik (cm)", "Gramaj (gr)", "Renk", "Kat", "Toplam Agirlik (kg)"]
        for col, h in enumerate(headers, 1):
            ws1.cell(row=3, column=col, value=h)
        style_header(ws1, 3, len(headers))
        for idx, b in enumerate(bobins):
            row = 4 + idx
            ws1.cell(row=row, column=1, value=b.get("barcode", ""))
            ws1.cell(row=row, column=2, value=b.get("brand", ""))
            ws1.cell(row=row, column=3, value=b.get("width_cm", 0)).alignment = center_align
            ws1.cell(row=row, column=4, value=b.get("grammage", 0)).alignment = center_align
            ws1.cell(row=row, column=5, value=b.get("color", ""))
            ws1.cell(row=row, column=6, value=_layer_label(b)).alignment = center_align
            ws1.cell(row=row, column=7, value=round(b.get("total_weight_kg", 0), 2)).alignment = center_align
        total_row = 4 + len(bobins)
        ws1.cell(row=total_row, column=1, value="TOPLAM").font = bold
        ws1.cell(row=total_row, column=7,
                 value=round(sum(b.get("total_weight_kg", 0) for b in bobins), 2)).alignment = center_align
        ws1.cell(row=total_row, column=7).font = bold
        auto_width(ws1, [22, 18, 14, 14, 14, 12, 22])

    # ==================== SAYFA 2: TÜM HAREKETLER ====================
    ws2 = wb.create_sheet("Hareketler")
    title_row(ws2, 7, f"Bobin Hareket Gecmisi {('— ' + period_label) if is_archive else ''}")
    type_map = {"purchase": "Satin Alma", "to_machine": "Makineye", "sale": "Satis", "adjustment": "Duzeltme"}
    h2 = ["Tarih", "Bobin", "Islem", "Agirlik (kg)", "Makine/Musteri", "Kullanici", "Not"]
    for col, h in enumerate(h2, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header(ws2, 3, len(h2))
    for idx, m in enumerate(movements):
        row = 4 + idx
        created = m.get("created_at", "")
        try:
            dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
            date_str = dt.strftime("%d.%m.%Y %H:%M")
        except Exception:
            date_str = created[:16] if created else "-"
        prefix = "+" if m.get("movement_type") == "purchase" else "-"
        ws2.cell(row=row, column=1, value=date_str)
        ws2.cell(row=row, column=2, value=m.get("bobin_label", ""))
        ws2.cell(row=row, column=3, value=type_map.get(m.get("movement_type", ""), m.get("movement_type", "")))
        ws2.cell(row=row, column=4, value=f"{prefix}{round(m.get('weight_kg', 0), 2)}").alignment = center_align
        ws2.cell(row=row, column=5, value=m.get("machine_name", "") or m.get("customer_name", "") or "-")
        ws2.cell(row=row, column=6, value=m.get("user_name", ""))
        ws2.cell(row=row, column=7, value=m.get("note", ""))
    auto_width(ws2, [18, 38, 14, 14, 22, 16, 24])

    # ==================== SAYFA 3: MAKİNEYE DAĞILIM ====================
    ws3 = wb.create_sheet("Makine Dagilimi")
    title_row(ws3, 4, f"Makine Bazli Bobin Tuketimi {('— ' + period_label) if is_archive else '(Tum Zamanlar)'}")
    machine_totals = {}  # machine_name -> {"kg": x, "count": n}
    for m in movements:
        if m.get("movement_type") == "to_machine":
            mn = m.get("machine_name", "-") or "-"
            entry = machine_totals.setdefault(mn, {"kg": 0, "count": 0})
            entry["kg"] += float(m.get("weight_kg", 0) or 0)
            entry["count"] += 1
    h3 = ["Makine", "Toplam Agirlik (kg)", "Hareket Sayisi", "Ortalama (kg/hareket)"]
    for col, h in enumerate(h3, 1):
        ws3.cell(row=3, column=col, value=h)
    style_header(ws3, 3, len(h3))
    sorted_m = sorted(machine_totals.items(), key=lambda x: -x[1]["kg"])
    for idx, (mn, info) in enumerate(sorted_m):
        row = 4 + idx
        avg = round(info["kg"] / info["count"], 2) if info["count"] else 0
        ws3.cell(row=row, column=1, value=mn)
        ws3.cell(row=row, column=2, value=round(info["kg"], 2)).alignment = center_align
        ws3.cell(row=row, column=3, value=info["count"]).alignment = center_align
        ws3.cell(row=row, column=4, value=avg).alignment = center_align
    if sorted_m:
        total_r = 4 + len(sorted_m)
        ws3.cell(row=total_r, column=1, value="TOPLAM").font = bold
        ws3.cell(row=total_r, column=2, value=round(sum(i["kg"] for _, i in sorted_m), 2)).alignment = center_align
        ws3.cell(row=total_r, column=2).font = bold
        ws3.cell(row=total_r, column=3, value=sum(i["count"] for _, i in sorted_m)).alignment = center_align
        ws3.cell(row=total_r, column=3).font = bold
    auto_width(ws3, [28, 22, 16, 22])

    # ==================== SAYFA 4: MÜŞTERİ SATIŞLARI ====================
    ws4 = wb.create_sheet("Musteri Satislari")
    title_row(ws4, 4, f"Musteri Bazli Bobin Satislari {('— ' + period_label) if is_archive else '(Tum Zamanlar)'}")
    customer_totals = {}
    for m in movements:
        if m.get("movement_type") == "sale":
            cn = m.get("customer_name", "-") or "-"
            entry = customer_totals.setdefault(cn, {"kg": 0, "count": 0})
            entry["kg"] += float(m.get("weight_kg", 0) or 0)
            entry["count"] += 1
    h4 = ["Musteri", "Toplam Agirlik (kg)", "Satis Sayisi", "Ortalama (kg/satis)"]
    for col, h in enumerate(h4, 1):
        ws4.cell(row=3, column=col, value=h)
    style_header(ws4, 3, len(h4))
    sorted_c = sorted(customer_totals.items(), key=lambda x: -x[1]["kg"])
    for idx, (cn, info) in enumerate(sorted_c):
        row = 4 + idx
        avg = round(info["kg"] / info["count"], 2) if info["count"] else 0
        ws4.cell(row=row, column=1, value=cn)
        ws4.cell(row=row, column=2, value=round(info["kg"], 2)).alignment = center_align
        ws4.cell(row=row, column=3, value=info["count"]).alignment = center_align
        ws4.cell(row=row, column=4, value=avg).alignment = center_align
    if sorted_c:
        total_r = 4 + len(sorted_c)
        ws4.cell(row=total_r, column=1, value="TOPLAM").font = bold
        ws4.cell(row=total_r, column=2, value=round(sum(i["kg"] for _, i in sorted_c), 2)).alignment = center_align
        ws4.cell(row=total_r, column=2).font = bold
    auto_width(ws4, [28, 22, 16, 22])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    if is_archive:
        await log_audit(user_display, "export_archive", "bobin", period_label,
                        f"Aylık arşiv indirildi ({len(movements)} hareket)")
        filename = f"bobin_arsiv_{month}.xlsx"
    else:
        await log_audit(user_display, "export", "bobin", "Excel",
                        f"Bobin stok raporu indirildi ({len(bobins)} tur, {len(movements)} hareket)")
        filename = f"bobin_stok_{datetime.now(timezone.utc).strftime('%d%m%Y_%H%M')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
