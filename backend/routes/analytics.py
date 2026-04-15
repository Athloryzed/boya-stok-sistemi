from fastapi import APIRouter
from typing import Optional
from datetime import datetime, timezone, timedelta
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import db

router = APIRouter()


@router.get("/analytics/weekly")
async def get_weekly_analytics():
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    week_ago_str = week_ago.isoformat()

    completed_jobs = await db.jobs.find(
        {"status": "completed", "completed_at": {"$gte": week_ago_str}}, {"_id": 0}
    ).to_list(1000)

    shift_reports = await db.shift_end_reports.find(
        {"created_at": {"$gte": week_ago_str}}, {"_id": 0}
    ).to_list(1000)

    machine_stats = {}
    for job in completed_jobs:
        machine = job["machine_name"]
        koli = job.get("completed_koli", job.get("koli_count", 0))
        if machine not in machine_stats:
            machine_stats[machine] = 0
        machine_stats[machine] += koli

    job_ids = [r["job_id"] for r in shift_reports if r.get("job_id")]
    if job_ids:
        jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
        jobs_dict = {j["id"]: j for j in jobs_list}
        for report in shift_reports:
            job_id = report.get("job_id")
            if job_id:
                job = jobs_dict.get(job_id)
                if job and job.get("status") != "completed":
                    machine = report.get("machine_name", "")
                    koli = report.get("produced_koli", 0)
                    if machine and koli > 0:
                        if machine not in machine_stats:
                            machine_stats[machine] = 0
                        machine_stats[machine] += koli

    return {"machine_stats": machine_stats}


@router.get("/analytics/daily")
async def get_daily_analytics():
    daily_stats = []
    for i in range(7):
        date = datetime.now(timezone.utc) - timedelta(days=i)
        start_of_day = date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)

        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)

        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)

        total_koli = sum(job.get("completed_koli", job.get("koli_count", 0)) for job in jobs)
        machine_breakdown = {}

        for job in jobs:
            machine = job["machine_name"]
            if machine not in machine_breakdown:
                machine_breakdown[machine] = 0
            machine_breakdown[machine] += job.get("completed_koli", job.get("koli_count", 0))

        job_ids = [r["job_id"] for r in shift_reports if r.get("job_id")]
        if job_ids:
            jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
            jobs_dict = {j["id"]: j for j in jobs_list}
            for report in shift_reports:
                job_id = report.get("job_id")
                if job_id:
                    job = jobs_dict.get(job_id)
                    if job and job.get("status") != "completed":
                        machine = report.get("machine_name", "")
                        koli = report.get("produced_koli", 0)
                        if machine and koli > 0:
                            if machine not in machine_breakdown:
                                machine_breakdown[machine] = 0
                            machine_breakdown[machine] += koli
                            total_koli += koli

        daily_stats.append({
            "date": start_of_day.strftime("%d %b"),
            "total_koli": total_koli,
            "machines": machine_breakdown
        })

    return {"daily_stats": list(reversed(daily_stats))}


@router.get("/analytics/daily-detail")
async def get_daily_detail_analytics(date: str):
    """Belirli bir günün detaylı üretim analitiği"""
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError:
        return {"error": "Geçersiz tarih formatı. YYYY-MM-DD kullanın."}

    start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = start_of_day + timedelta(days=1)

    completed_jobs = await db.jobs.find(
        {"status": "completed", "completed_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
        {"_id": 0}
    ).to_list(1000)

    started_jobs = await db.jobs.find(
        {"started_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
        {"_id": 0}
    ).to_list(1000)

    shift_reports = await db.shift_end_reports.find(
        {"created_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
        {"_id": 0}
    ).to_list(1000)

    defects = await db.defect_logs.find(
        {"created_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
        {"_id": 0}
    ).to_list(1000)

    machine_breakdown = {}
    operator_breakdown = {}
    job_details = []

    for job in completed_jobs:
        machine = job.get("machine_name", "Bilinmiyor")
        operator = job.get("operator_name", "Bilinmiyor")
        koli = job.get("completed_koli", job.get("koli_count", 0))

        if machine not in machine_breakdown:
            machine_breakdown[machine] = 0
        machine_breakdown[machine] += koli

        if operator not in operator_breakdown:
            operator_breakdown[operator] = {"jobs_count": 0, "total_koli": 0}
        operator_breakdown[operator]["jobs_count"] += 1
        operator_breakdown[operator]["total_koli"] += koli

        duration_min = None
        if job.get("started_at") and job.get("completed_at"):
            try:
                start = datetime.fromisoformat(job["started_at"].replace("Z", "+00:00"))
                end = datetime.fromisoformat(job["completed_at"].replace("Z", "+00:00"))
                duration_min = round((end - start).total_seconds() / 60)
            except Exception:
                pass

        job_details.append({
            "id": job.get("id"), "name": job.get("name", ""),
            "machine_name": machine, "operator_name": operator,
            "koli_count": koli, "started_at": job.get("started_at"),
            "completed_at": job.get("completed_at"),
            "duration_min": duration_min, "colors": job.get("colors", "")
        })

    partial_koli = 0
    if shift_reports:
        job_ids = [r["job_id"] for r in shift_reports if r.get("job_id") and r.get("produced_koli", 0) > 0]
        if job_ids:
            jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
            jobs_dict = {j["id"]: j for j in jobs_list}
            for report in shift_reports:
                produced = report.get("produced_koli", 0)
                if produced > 0:
                    job_id = report.get("job_id")
                    if job_id:
                        job = jobs_dict.get(job_id)
                        if job and job.get("status") == "completed":
                            continue
                    machine = report.get("machine_name", "")
                    if machine:
                        if machine not in machine_breakdown:
                            machine_breakdown[machine] = 0
                        machine_breakdown[machine] += produced
                        partial_koli += produced

    total_defect_kg = 0.0
    defect_by_machine = {}
    for defect in defects:
        kg = defect.get("defect_kg", 0)
        total_defect_kg += kg
        machine = defect.get("machine_name", "Bilinmiyor")
        if machine not in defect_by_machine:
            defect_by_machine[machine] = 0.0
        defect_by_machine[machine] += kg

    total_koli = sum(machine_breakdown.values())
    machine_chart = [{"name": k, "koli": v} for k, v in sorted(machine_breakdown.items(), key=lambda x: x[1], reverse=True)]
    operator_chart = [{"name": k, "koli": v["total_koli"], "jobs": v["jobs_count"]} for k, v in sorted(operator_breakdown.items(), key=lambda x: x[1]["total_koli"], reverse=True)]

    return {
        "date": date,
        "summary": {
            "total_koli": total_koli,
            "completed_jobs": len(completed_jobs),
            "started_jobs": len(started_jobs),
            "partial_koli": partial_koli,
            "active_operators": len(operator_breakdown),
            "total_defect_kg": round(total_defect_kg, 2)
        },
        "machine_chart": machine_chart,
        "operator_chart": operator_chart,
        "job_details": sorted(job_details, key=lambda x: x.get("completed_at", ""), reverse=True),
        "defect_by_machine": {k: round(v, 2) for k, v in defect_by_machine.items()}
    }


@router.get("/analytics/monthly")
async def get_monthly_analytics(year: Optional[int] = None, month: Optional[int] = None):
    if year and month:
        if month == 12:
            start_date = datetime(year, month, 1, tzinfo=timezone.utc)
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            start_date = datetime(year, month, 1, tzinfo=timezone.utc)
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        start_date_str = start_date.isoformat()
        end_date_str = end_date.isoformat()

        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": start_date_str, "$lt": end_date_str}},
            {"_id": 0}
        ).to_list(1000)

        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": start_date_str, "$lt": end_date_str}},
            {"_id": 0}
        ).to_list(1000)
    else:
        month_ago = datetime.now(timezone.utc) - timedelta(days=30)
        month_ago_str = month_ago.isoformat()

        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": month_ago_str}},
            {"_id": 0}
        ).to_list(1000)

        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": month_ago_str}},
            {"_id": 0}
        ).to_list(1000)

    machine_stats = {}
    for job in jobs:
        machine = job["machine_name"]
        koli = job.get("completed_koli", job.get("koli_count", 0))
        if machine not in machine_stats:
            machine_stats[machine] = 0
        machine_stats[machine] += koli

    job_ids = [r["job_id"] for r in shift_reports if r.get("job_id")]
    if job_ids:
        jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
        jobs_dict = {j["id"]: j for j in jobs_list}
        for report in shift_reports:
            job_id = report.get("job_id")
            if job_id:
                job = jobs_dict.get(job_id)
                if job and job.get("status") != "completed":
                    machine = report.get("machine_name", "")
                    koli = report.get("produced_koli", 0)
                    if machine and koli > 0:
                        if machine not in machine_stats:
                            machine_stats[machine] = 0
                        machine_stats[machine] += koli

    return {"machine_stats": machine_stats}


@router.get("/analytics/daily-by-week")
async def get_daily_analytics_by_week(week_offset: int = 0):
    """Hafta bazında günlük üretim analitiği"""
    today = datetime.now(timezone.utc)
    days_since_monday = today.weekday()
    this_monday = today - timedelta(days=days_since_monday)
    target_monday = this_monday + timedelta(weeks=week_offset)
    target_sunday = target_monday + timedelta(days=6)

    daily_stats = []
    for i in range(7):
        date = target_monday + timedelta(days=i)
        start_of_day = date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)

        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)

        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)

        total_koli = sum(job.get("completed_koli", job.get("koli_count", 0)) for job in jobs)
        machine_breakdown = {}

        for job in jobs:
            machine = job["machine_name"]
            if machine not in machine_breakdown:
                machine_breakdown[machine] = 0
            machine_breakdown[machine] += job.get("completed_koli", job.get("koli_count", 0))

        job_ids = [r["job_id"] for r in shift_reports if r.get("job_id") and r.get("produced_koli", 0) > 0]
        if job_ids:
            jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
            jobs_dict = {j["id"]: j for j in jobs_list}
            for report in shift_reports:
                produced = report.get("produced_koli", 0)
                if produced > 0:
                    job_id = report.get("job_id")
                    if job_id:
                        job = jobs_dict.get(job_id)
                        if job and job.get("status") == "completed":
                            continue
                    machine = report.get("machine_name", "")
                    if machine:
                        if machine not in machine_breakdown:
                            machine_breakdown[machine] = 0
                        machine_breakdown[machine] += produced
                        total_koli += produced

        day_names = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        daily_stats.append({
            "date": start_of_day.strftime("%d %b"),
            "day_name": day_names[i],
            "full_date": start_of_day.strftime("%Y-%m-%d"),
            "total_koli": total_koli,
            "machines": machine_breakdown
        })

    return {
        "week_start": target_monday.strftime("%d %b %Y"),
        "week_end": target_sunday.strftime("%d %b %Y"),
        "week_offset": week_offset,
        "daily_stats": daily_stats
    }


@router.get("/analytics/export")
async def export_analytics(period: str = "weekly", week_offset: int = 0):
    if period == "weekly":
        today = datetime.now(timezone.utc)
        days_since_monday = today.weekday()
        this_monday = today - timedelta(days=days_since_monday)
        target_monday = this_monday + timedelta(weeks=week_offset)
        target_sunday = target_monday + timedelta(days=6)
        start_date = target_monday.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = target_sunday.replace(hour=23, minute=59, second=59) + timedelta(seconds=1)
        period_label = f"{target_monday.strftime('%d.%m.%Y')} - {target_sunday.strftime('%d.%m.%Y')}"
    else:
        date_ago = datetime.now(timezone.utc) - timedelta(days=30)
        start_date = date_ago
        end_date = datetime.now(timezone.utc)
        period_label = f"Son 30 Gün ({date_ago.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')})"

    start_str = start_date.isoformat()
    end_str = end_date.isoformat()

    completed_jobs = await db.jobs.find(
        {"status": "completed", "completed_at": {"$gte": start_str, "$lt": end_str}}, {"_id": 0}
    ).to_list(1000)
    started_jobs = await db.jobs.find(
        {"started_at": {"$gte": start_str, "$lt": end_str}}, {"_id": 0}
    ).to_list(1000)
    shift_reports = await db.shift_end_reports.find(
        {"created_at": {"$gte": start_str, "$lt": end_str}}, {"_id": 0}
    ).to_list(1000)
    defects = await db.defect_logs.find(
        {"created_at": {"$gte": start_str, "$lt": end_str}}, {"_id": 0}
    ).to_list(1000)

    # Stiller
    wb = Workbook()
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    sub_header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

    def auto_width(ws, cols):
        for col in range(1, cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # SAYFA 1: HAFTALIK ÖZET
    ws1 = wb.active
    ws1.title = "Haftalik Ozet"
    ws1.merge_cells("A1:G1")
    title_cell = ws1.cell(row=1, column=1, value=f"BUSE KAGIT - Haftalik Ozet Raporu ({period_label})")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    ws1.row_dimensions[1].height = 30

    day_names_tr = ["Pazartesi", "Sali", "Carsamba", "Persembe", "Cuma", "Cumartesi", "Pazar"]
    headers1 = ["Gun", "Tarih", "Toplam Koli", "Tamamlanan Is", "Baslayan Is", "Operator Sayisi", "Defo (kg)"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=3, column=col, value=h)
    style_header(ws1, 3, 7)

    grand_total_koli = 0
    grand_total_completed = 0
    grand_total_started = 0
    all_operators = set()
    grand_total_defect = 0.0

    for i in range(7):
        day = start_date + timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)

        day_jobs = [j for j in completed_jobs if day_start.isoformat() <= j.get("completed_at", "") < day_end.isoformat()]
        day_started = [j for j in started_jobs if day_start.isoformat() <= j.get("started_at", "") < day_end.isoformat()]
        day_defects = [d for d in defects if day_start.isoformat() <= d.get("created_at", "") < day_end.isoformat()]

        day_koli = sum(j.get("completed_koli", j.get("koli_count", 0)) for j in day_jobs)
        day_operators = set(j.get("operator_name", "") for j in day_jobs if j.get("operator_name"))
        day_defect_kg = sum(d.get("defect_kg", 0) for d in day_defects)

        grand_total_koli += day_koli
        grand_total_completed += len(day_jobs)
        grand_total_started += len(day_started)
        all_operators.update(day_operators)
        grand_total_defect += day_defect_kg

        row = 4 + i
        ws1.cell(row=row, column=1, value=day_names_tr[i])
        ws1.cell(row=row, column=2, value=day.strftime("%d.%m.%Y"))
        ws1.cell(row=row, column=3, value=day_koli).alignment = center_align
        ws1.cell(row=row, column=4, value=len(day_jobs)).alignment = center_align
        ws1.cell(row=row, column=5, value=len(day_started)).alignment = center_align
        ws1.cell(row=row, column=6, value=len(day_operators)).alignment = center_align
        ws1.cell(row=row, column=7, value=round(day_defect_kg, 2)).alignment = center_align

    total_row = 11
    for col in range(1, 8):
        ws1.cell(row=total_row, column=col).fill = sub_header_fill
        ws1.cell(row=total_row, column=col).font = Font(bold=True)
    ws1.cell(row=total_row, column=1, value="TOPLAM")
    ws1.cell(row=total_row, column=3, value=grand_total_koli).alignment = center_align
    ws1.cell(row=total_row, column=4, value=grand_total_completed).alignment = center_align
    ws1.cell(row=total_row, column=5, value=grand_total_started).alignment = center_align
    ws1.cell(row=total_row, column=6, value=len(all_operators)).alignment = center_align
    ws1.cell(row=total_row, column=7, value=round(grand_total_defect, 2)).alignment = center_align
    auto_width(ws1, 7)

    # SAYFA 2: MAKİNE DETAYI
    ws2 = wb.create_sheet("Makine Detayi")
    ws2.merge_cells("A1:H1")
    title2 = ws2.cell(row=1, column=1, value=f"Makine Bazinda Uretim Detayi ({period_label})")
    title2.font = title_font
    title2.fill = title_fill
    title2.alignment = center_align
    ws2.row_dimensions[1].height = 30

    all_machines = sorted(set(j.get("machine_name", "") for j in completed_jobs if j.get("machine_name")))
    headers2 = ["Makine"] + day_names_tr + ["TOPLAM"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header(ws2, 3, len(headers2))

    for m_idx, machine in enumerate(all_machines):
        row = 4 + m_idx
        ws2.cell(row=row, column=1, value=machine)
        machine_total = 0
        for d_idx in range(7):
            day = start_date + timedelta(days=d_idx)
            day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            day_machine_koli = sum(
                j.get("completed_koli", j.get("koli_count", 0))
                for j in completed_jobs
                if j.get("machine_name") == machine and day_start.isoformat() <= j.get("completed_at", "") < day_end.isoformat()
            )
            ws2.cell(row=row, column=2 + d_idx, value=day_machine_koli).alignment = center_align
            machine_total += day_machine_koli
        ws2.cell(row=row, column=9, value=machine_total).alignment = center_align
        ws2.cell(row=row, column=9).font = Font(bold=True)
    auto_width(ws2, 9)

    # SAYFA 3: OPERATÖR PERFORMANSI
    ws3 = wb.create_sheet("Operator Performansi")
    ws3.merge_cells("A1:F1")
    title3 = ws3.cell(row=1, column=1, value=f"Operator Performansi ({period_label})")
    title3.font = title_font
    title3.fill = title_fill
    title3.alignment = center_align
    ws3.row_dimensions[1].height = 30

    headers3 = ["Operator", "Tamamlanan Is", "Toplam Koli", "Ort. Koli/Is", "Calistigi Makineler", "Ort. Sure (dk)"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=col, value=h)
    style_header(ws3, 3, 6)

    op_stats = {}
    for job in completed_jobs:
        op = job.get("operator_name", "Bilinmiyor")
        if op not in op_stats:
            op_stats[op] = {"jobs": 0, "koli": 0, "machines": set(), "durations": []}
        op_stats[op]["jobs"] += 1
        op_stats[op]["koli"] += job.get("completed_koli", job.get("koli_count", 0))
        op_stats[op]["machines"].add(job.get("machine_name", ""))
        if job.get("started_at") and job.get("completed_at"):
            try:
                s = datetime.fromisoformat(job["started_at"].replace("Z", "+00:00"))
                e = datetime.fromisoformat(job["completed_at"].replace("Z", "+00:00"))
                op_stats[op]["durations"].append((e - s).total_seconds() / 60)
            except Exception:
                pass

    for idx, (op, stats) in enumerate(sorted(op_stats.items(), key=lambda x: x[1]["koli"], reverse=True)):
        row = 4 + idx
        avg_koli = round(stats["koli"] / stats["jobs"]) if stats["jobs"] > 0 else 0
        avg_dur = round(sum(stats["durations"]) / len(stats["durations"])) if stats["durations"] else 0
        ws3.cell(row=row, column=1, value=op)
        ws3.cell(row=row, column=2, value=stats["jobs"]).alignment = center_align
        ws3.cell(row=row, column=3, value=stats["koli"]).alignment = center_align
        ws3.cell(row=row, column=4, value=avg_koli).alignment = center_align
        ws3.cell(row=row, column=5, value=", ".join(sorted(stats["machines"])))
        ws3.cell(row=row, column=6, value=avg_dur).alignment = center_align
    auto_width(ws3, 6)

    # SAYFA 4: DEFO RAPORU
    ws4 = wb.create_sheet("Defo Raporu")
    ws4.merge_cells("A1:E1")
    title4 = ws4.cell(row=1, column=1, value=f"Defo Raporu ({period_label})")
    title4.font = title_font
    title4.fill = title_fill
    title4.alignment = center_align
    ws4.row_dimensions[1].height = 30

    headers4 = ["Tarih", "Makine", "Defo (kg)", "Operator", "Aciklama"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=col, value=h)
    style_header(ws4, 3, 5)

    sorted_defects = sorted(defects, key=lambda x: x.get("created_at", ""), reverse=True)
    for idx, defect in enumerate(sorted_defects):
        row = 4 + idx
        created = defect.get("created_at", "")
        try:
            dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
            date_str = dt.strftime("%d.%m.%Y %H:%M")
        except Exception:
            date_str = created[:16] if created else "-"
        ws4.cell(row=row, column=1, value=date_str)
        ws4.cell(row=row, column=2, value=defect.get("machine_name", "-"))
        ws4.cell(row=row, column=3, value=round(defect.get("defect_kg", 0), 2)).alignment = center_align
        ws4.cell(row=row, column=4, value=defect.get("operator_name", "-"))
        ws4.cell(row=row, column=5, value=defect.get("notes", "-"))

    if not sorted_defects:
        ws4.cell(row=4, column=1, value="Bu donemde defo kaydi bulunmamaktadir.")

    if sorted_defects:
        summary_row = 4 + len(sorted_defects) + 1
        ws4.cell(row=summary_row, column=1, value="TOPLAM").font = Font(bold=True)
        ws4.cell(row=summary_row, column=3, value=round(sum(d.get("defect_kg", 0) for d in defects), 2))
        ws4.cell(row=summary_row, column=3).font = Font(bold=True)
        ws4.cell(row=summary_row, column=3).alignment = center_align
        for col in range(1, 6):
            ws4.cell(row=summary_row, column=col).fill = sub_header_fill

    auto_width(ws4, 5)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"buse_kagit_rapor_{period}_{start_date.strftime('%d%m%Y')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
