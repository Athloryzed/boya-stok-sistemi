"""
Koli Stoğu (Box Stock) — Hammadde/sarf koli takibi.
Fabrikaya alınan koliler stoğa girer; makinelere kullanılması için verilir.
"""
from fastapi import APIRouter, HTTPException, Body, Depends
from typing import Optional
from datetime import datetime, timezone, timedelta
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from database import db
from models import KoliStock, KoliStockMovement
from auth import get_current_user
from services.audit import log_audit

router = APIRouter(dependencies=[Depends(get_current_user)])
logger = logging.getLogger(__name__)


def koli_label(s: dict) -> str:
    parts = [s.get("name", "")]
    if s.get("size"):
        parts.append(s["size"])
    return " - ".join([p for p in parts if p])


@router.get("/koli-stock")
async def list_koli(name: Optional[str] = None, size: Optional[str] = None):
    query: dict = {}
    if name:
        query["name"] = name
    if size:
        query["size"] = size
    items = await db.koli_stock.find(query, {"_id": 0}).sort([("name", 1)]).to_list(500)
    return items


@router.post("/koli-stock")
async def add_koli(data: dict = Body(...)):
    """Stoğa koli ekle. Aynı isim+ölçü varsa merge eder."""
    name = (data.get("name") or "").strip()
    size = (data.get("size") or "").strip() or None
    quantity = int(data.get("quantity", 0) or 0)
    notes = (data.get("notes") or "").strip() or None
    supplier = (data.get("supplier") or "").strip() or None
    user_name = (data.get("user_name") or "").strip() or "Depo"

    if not name:
        raise HTTPException(status_code=400, detail="Koli adı zorunludur")
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Adet sıfırdan büyük olmalıdır")

    query: dict = {"name": name}
    if size:
        query["size"] = size
    else:
        query["size"] = {"$in": [None, ""]}

    existing = await db.koli_stock.find_one(query, {"_id": 0})
    now = datetime.now(timezone.utc).isoformat()

    if existing:
        new_qty = int(existing.get("quantity", 0)) + quantity
        await db.koli_stock.update_one(
            {"id": existing["id"]},
            {"$set": {
                "quantity": new_qty,
                "updated_at": now,
                **({"notes": notes} if notes else {}),
            }}
        )
        stock_id = existing["id"]
        stock_doc = {**existing, "quantity": new_qty, "notes": notes or existing.get("notes")}
    else:
        stock = KoliStock(name=name, size=size, quantity=quantity, notes=notes)
        stock_doc = stock.model_dump()
        await db.koli_stock.insert_one(stock_doc)
        stock_id = stock.id
        stock_doc.pop("_id", None)

    mv = KoliStockMovement(
        stock_id=stock_id, name=name, size=size,
        movement_type="in", quantity=quantity,
        supplier=supplier, note=notes, user_name=user_name
    )
    await db.koli_stock_movements.insert_one(mv.model_dump())

    await log_audit(
        user_name, "create", "koli_stock", stock_id,
        f"Koli stoğa eklendi: {koli_label({'name': name, 'size': size})} +{quantity} adet"
        + (f" (Tedarikçi: {supplier})" if supplier else "")
    )

    return {"success": True, "stock": stock_doc, "movement_id": mv.id}


@router.post("/koli-stock/give-to-machine")
async def give_to_machine(data: dict = Body(...)):
    """Makineye koli ver — stoktan düşer, hareket olarak kaydedilir."""
    stock_id = (data.get("stock_id") or "").strip()
    quantity = int(data.get("quantity", 0) or 0)
    machine_id = (data.get("machine_id") or "").strip() or None
    machine_name = (data.get("machine_name") or "").strip() or None
    note = (data.get("note") or "").strip() or None
    user_name = (data.get("user_name") or "").strip() or "Depo"

    if not stock_id:
        raise HTTPException(status_code=400, detail="Koli seçilmedi")
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Adet sıfırdan büyük olmalıdır")

    stock = await db.koli_stock.find_one({"id": stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Koli stoğu bulunamadı")

    current = int(stock.get("quantity", 0))
    if quantity > current:
        raise HTTPException(status_code=400, detail=f"Yetersiz stok (mevcut: {current})")

    # Makine adını id'den çek (verilmediyse)
    if machine_id and not machine_name:
        m = await db.machines.find_one({"id": machine_id}, {"_id": 0, "name": 1})
        if m:
            machine_name = m.get("name")

    new_qty = current - quantity
    now = datetime.now(timezone.utc).isoformat()
    await db.koli_stock.update_one(
        {"id": stock_id}, {"$set": {"quantity": new_qty, "updated_at": now}}
    )

    mv = KoliStockMovement(
        stock_id=stock_id, name=stock["name"], size=stock.get("size"),
        movement_type="out", quantity=quantity,
        machine_id=machine_id, machine_name=machine_name,
        note=note, user_name=user_name
    )
    await db.koli_stock_movements.insert_one(mv.model_dump())

    await log_audit(
        user_name, "update", "koli_stock", stock_id,
        f"Makineye verildi: {koli_label(stock)} -{quantity} adet"
        + (f" → {machine_name}" if machine_name else "")
    )

    return {"success": True, "new_quantity": new_qty, "movement_id": mv.id}


@router.patch("/koli-stock/{stock_id}")
async def edit_koli(stock_id: str, data: dict = Body(...)):
    user_name = (data.get("user_name") or "").strip() or "Depo"
    stock = await db.koli_stock.find_one({"id": stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Koli stoğu bulunamadı")

    updates = {}
    for fld in ("name", "size", "notes"):
        if fld in data:
            v = (data.get(fld) or "")
            if isinstance(v, str):
                v = v.strip() or None
            updates[fld] = v
    if "quantity" in data:
        new_q = int(data.get("quantity", 0) or 0)
        if new_q < 0:
            raise HTTPException(status_code=400, detail="Adet negatif olamaz")
        delta = new_q - int(stock.get("quantity", 0))
        updates["quantity"] = new_q
        if delta != 0:
            mv = KoliStockMovement(
                stock_id=stock_id, name=stock["name"], size=stock.get("size"),
                movement_type="adjustment", quantity=abs(delta),
                note=f"Düzeltme: {int(stock.get('quantity', 0))} → {new_q}",
                user_name=user_name
            )
            await db.koli_stock_movements.insert_one(mv.model_dump())

    if not updates:
        return {"success": True, "message": "Değişiklik yok"}

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.koli_stock.update_one({"id": stock_id}, {"$set": updates})

    await log_audit(
        user_name, "update", "koli_stock", stock_id,
        f"Düzeltme: {koli_label(stock)} — {', '.join([f'{k}={v}' for k, v in updates.items() if k != 'updated_at'])}"
    )
    new_stock = await db.koli_stock.find_one({"id": stock_id}, {"_id": 0})
    return {"success": True, "stock": new_stock}


@router.delete("/koli-stock/{stock_id}")
async def delete_koli(stock_id: str, data: dict = Body(None)):
    user_name = (data.get("user_name") if data else None) or "Depo"
    stock = await db.koli_stock.find_one({"id": stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Koli stoğu bulunamadı")
    await db.koli_stock.delete_one({"id": stock_id})
    await log_audit(
        user_name, "delete", "koli_stock", stock_id,
        f"Silindi: {koli_label(stock)} (son adet: {stock.get('quantity', 0)})"
    )
    return {"success": True}


@router.get("/koli-stock/movements")
async def list_movements(
    stock_id: Optional[str] = None,
    movement_type: Optional[str] = None,
    limit: int = 200,
):
    query: dict = {}
    if stock_id:
        query["stock_id"] = stock_id
    if movement_type:
        query["movement_type"] = movement_type
    return await db.koli_stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)


@router.get("/koli-stock/summary")
async def koli_summary(days: int = 30):
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    pipeline = [
        {"$match": {"created_at": {"$gte": cutoff}}},
        {"$group": {
            "_id": {"name": "$name", "type": "$movement_type"},
            "total": {"$sum": "$quantity"}
        }},
    ]
    rows = await db.koli_stock_movements.aggregate(pipeline).to_list(1000)
    by_name: dict = {}
    for r in rows:
        n = r["_id"]["name"]
        t = r["_id"]["type"]
        if n not in by_name:
            by_name[n] = {"name": n, "in": 0, "out": 0, "adjustment": 0}
        by_name[n][t] = by_name[n].get(t, 0) + r["total"]

    current_stock = await db.koli_stock.find({}, {"_id": 0, "name": 1, "quantity": 1}).to_list(500)
    current_total = sum(int(s.get("quantity", 0)) for s in current_stock)

    return {
        "days": days,
        "rows": list(by_name.values()),
        "current_total": current_total,
        "current_count": len(current_stock),
    }


@router.get("/koli-stock/export")
async def export_koli():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Koli Stoğu"
    headers1 = ["Ad", "Ölçü", "Adet", "Not", "Oluşturma", "Güncelleme"]
    ws1.append(headers1)
    for col_idx, _ in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
    stocks = await db.koli_stock.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    for s in stocks:
        ws1.append([
            s.get("name", ""), s.get("size", "") or "",
            int(s.get("quantity", 0)), s.get("notes", "") or "",
            (s.get("created_at", "") or "")[:19].replace("T", " "),
            (s.get("updated_at", "") or "")[:19].replace("T", " "),
        ])
    for col_idx in range(1, len(headers1) + 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = 18

    ws2 = wb.create_sheet("Hareketler")
    headers2 = ["Tarih", "Ad", "Ölçü", "Tip", "Adet", "Makine", "Tedarikçi", "Kullanıcı", "Not"]
    ws2.append(headers2)
    for col_idx, _ in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
    movements = await db.koli_stock_movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    type_label = {"in": "Giriş", "out": "Makineye", "adjustment": "Düzeltme"}
    for m in movements:
        ws2.append([
            (m.get("created_at", "") or "")[:19].replace("T", " "),
            m.get("name", ""), m.get("size", "") or "",
            type_label.get(m.get("movement_type", ""), m.get("movement_type", "")),
            int(m.get("quantity", 0)),
            m.get("machine_name", "") or "",
            m.get("supplier", "") or "",
            m.get("user_name", "") or "",
            m.get("note", "") or "",
        ])
    for col_idx in range(1, len(headers2) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"koli_stok_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
