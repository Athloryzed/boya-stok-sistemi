"""
Marka Stoğu (Brand Stock) — Bitmiş ürün stok takibi.
Bobin sistemi raw kağıt rolleri için; bu modül bitmiş üretim ürünleri için (Deniz 33, Banko, vs).
"""
from fastapi import APIRouter, HTTPException, Body, Depends
from typing import Optional
from datetime import datetime, timezone, timedelta
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from database import db
from models import BrandStock, BrandStockMovement
from auth import get_current_user
from services.audit import log_audit

router = APIRouter(dependencies=[Depends(get_current_user)])
logger = logging.getLogger(__name__)


# Marka şablonları — frontend bu listeyi tüketir
BRAND_TEMPLATES = [
    {
        "brand": "Deniz 33",
        "machines": ["33 ICM", "SİES"],
        "color_required": False,
    },
    {
        "brand": "Banko",
        "machines": ["ICM", "Büyük Makine"],
        "color_required": False,  # opsiyonel ama varsa serbest metin
    },
]


def stock_label(s: dict) -> str:
    parts = [s.get("brand", ""), s.get("machine", "")]
    if s.get("color"):
        parts.append(s["color"])
    return " - ".join([p for p in parts if p])


@router.get("/brand-stock/templates")
async def get_brand_templates():
    """Marka şablonları (Deniz 33 / Banko + makine seçenekleri)."""
    return {"templates": BRAND_TEMPLATES}


@router.get("/brand-stock")
async def list_brand_stock(
    brand: Optional[str] = None,
    machine: Optional[str] = None,
    color: Optional[str] = None,
):
    """Mevcut stok listesi — opsiyonel filtreler."""
    query: dict = {}
    if brand:
        query["brand"] = brand
    if machine:
        query["machine"] = machine
    if color:
        query["color"] = color
    items = await db.brand_stock.find(query, {"_id": 0}).sort([("brand", 1), ("machine", 1)]).to_list(500)
    return items


@router.post("/brand-stock")
async def add_brand_stock(data: dict = Body(...)):
    """Stoğa ekle. Aynı marka+makine+renk varsa miktarlar birleştirilir."""
    brand = (data.get("brand") or "").strip()
    machine = (data.get("machine") or "").strip()
    color = (data.get("color") or "").strip() or None
    quantity = int(data.get("quantity", 0) or 0)
    notes = (data.get("notes") or "").strip() or None
    user_name = (data.get("user_name") or "").strip() or "Depo"

    if not brand:
        raise HTTPException(status_code=400, detail="Marka zorunludur")
    # Custom marka için makine opsiyonel; bos string'i None'a cevir
    machine_val = machine if machine else None
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Adet sıfırdan büyük olmalıdır")

    # Mevcut kayıt var mı?
    query: dict = {"brand": brand}
    if machine_val:
        query["machine"] = machine_val
    else:
        query["machine"] = {"$in": [None, ""]}
    if color:
        query["color"] = color
    else:
        query["color"] = {"$in": [None, ""]}

    existing = await db.brand_stock.find_one(query, {"_id": 0})

    now = datetime.now(timezone.utc).isoformat()
    if existing:
        new_qty = int(existing.get("quantity", 0)) + quantity
        await db.brand_stock.update_one(
            {"id": existing["id"]},
            {"$set": {
                "quantity": new_qty,
                "updated_at": now,
                **({"notes": notes} if notes else {}),
            }}
        )
        stock_id = existing["id"]
        stock_doc = {**existing, "quantity": new_qty, "notes": notes or existing.get("notes")}
    else:
        stock = BrandStock(
            brand=brand, machine=machine_val, color=color,
            quantity=quantity, notes=notes
        )
        stock_doc = stock.model_dump()
        await db.brand_stock.insert_one(stock_doc)
        stock_id = stock.id
        stock_doc.pop("_id", None)

    # Hareket logu
    movement = BrandStockMovement(
        stock_id=stock_id, brand=brand, machine=machine_val, color=color,
        movement_type="in", quantity=quantity, note=notes, user_name=user_name
    )
    await db.brand_stock_movements.insert_one(movement.model_dump())

    await log_audit(
        user_name, "create", "brand_stock", stock_id,
        f"Stoğa eklendi: {stock_label({'brand': brand, 'machine': machine_val, 'color': color})} +{quantity} adet"
    )

    return {"success": True, "stock": stock_doc, "movement_id": movement.id}


@router.post("/brand-stock/sell")
async def sell_brand_stock(data: dict = Body(...)):
    """Satış/çıkış — stoğu azaltır."""
    stock_id = (data.get("stock_id") or "").strip()
    quantity = int(data.get("quantity", 0) or 0)
    customer_name = (data.get("customer_name") or "").strip() or None
    note = (data.get("note") or "").strip() or None
    user_name = (data.get("user_name") or "").strip() or "Depo"

    if not stock_id:
        raise HTTPException(status_code=400, detail="Stok kaydı seçilmedi")
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Adet sıfırdan büyük olmalıdır")

    stock = await db.brand_stock.find_one({"id": stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stok bulunamadı")

    current = int(stock.get("quantity", 0))
    if quantity > current:
        raise HTTPException(status_code=400, detail=f"Yetersiz stok (mevcut: {current})")

    new_qty = current - quantity
    now = datetime.now(timezone.utc).isoformat()
    await db.brand_stock.update_one(
        {"id": stock_id},
        {"$set": {"quantity": new_qty, "updated_at": now}}
    )

    movement = BrandStockMovement(
        stock_id=stock_id, brand=stock["brand"], machine=stock["machine"],
        color=stock.get("color"),
        movement_type="out", quantity=quantity,
        customer_name=customer_name, note=note, user_name=user_name
    )
    await db.brand_stock_movements.insert_one(movement.model_dump())

    await log_audit(
        user_name, "update", "brand_stock", stock_id,
        f"Satış: {stock_label(stock)} -{quantity} adet" + (f" → {customer_name}" if customer_name else "")
    )

    return {"success": True, "new_quantity": new_qty, "movement_id": movement.id}


@router.patch("/brand-stock/{stock_id}")
async def edit_brand_stock(stock_id: str, data: dict = Body(...)):
    """Mevcut stoğu düzelt — yanlış girilen veriyi düzeltmek için. Hareket logu adjustment olarak yazılır."""
    user_name = (data.get("user_name") or "").strip() or "Depo"
    stock = await db.brand_stock.find_one({"id": stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stok bulunamadı")

    updates = {}
    for fld in ("brand", "machine", "color", "notes"):
        if fld in data:
            v = (data.get(fld) or "")
            if isinstance(v, str):
                v = v.strip() or None
            updates[fld] = v
    if "quantity" in data:
        new_q = int(data.get("quantity", 0) or 0)
        if new_q < 0:
            raise HTTPException(status_code=400, detail="Adet negatif olamaz")
        delta = new_q - int(stock.get("quantity", 0))
        updates["quantity"] = new_q
        if delta != 0:
            mv = BrandStockMovement(
                stock_id=stock_id, brand=stock["brand"], machine=stock["machine"],
                color=stock.get("color"),
                movement_type="adjustment", quantity=abs(delta),
                note=f"Düzeltme: {int(stock.get('quantity', 0))} → {new_q}",
                user_name=user_name
            )
            await db.brand_stock_movements.insert_one(mv.model_dump())

    if not updates:
        return {"success": True, "message": "Değişiklik yok"}

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.brand_stock.update_one({"id": stock_id}, {"$set": updates})

    await log_audit(
        user_name, "update", "brand_stock", stock_id,
        f"Düzeltme: {stock_label(stock)} — {', '.join([f'{k}={v}' for k, v in updates.items() if k != 'updated_at'])}"
    )

    new_stock = await db.brand_stock.find_one({"id": stock_id}, {"_id": 0})
    return {"success": True, "stock": new_stock}


@router.delete("/brand-stock/{stock_id}")
async def delete_brand_stock(stock_id: str, data: dict = Body(None)):
    """Stoğu sil — sadece miktar 0 ise veya admin tarafından."""
    user_name = (data.get("user_name") if data else None) or "Yönetim"
    stock = await db.brand_stock.find_one({"id": stock_id}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stok bulunamadı")

    await db.brand_stock.delete_one({"id": stock_id})
    await log_audit(
        user_name, "delete", "brand_stock", stock_id,
        f"Silindi: {stock_label(stock)} (son adet: {stock.get('quantity', 0)})"
    )
    return {"success": True}


@router.get("/brand-stock/movements")
async def list_movements(
    stock_id: Optional[str] = None,
    brand: Optional[str] = None,
    movement_type: Optional[str] = None,
    limit: int = 200,
):
    """Hareket geçmişi."""
    query: dict = {}
    if stock_id:
        query["stock_id"] = stock_id
    if brand:
        query["brand"] = brand
    if movement_type:
        query["movement_type"] = movement_type
    movements = await db.brand_stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return movements


@router.get("/brand-stock/summary")
async def brand_stock_summary(days: int = 30):
    """Analiz özeti — marka bazlı son N gün üretim/satış."""
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    pipeline = [
        {"$match": {"created_at": {"$gte": cutoff}}},
        {"$group": {
            "_id": {"brand": "$brand", "machine": "$machine", "type": "$movement_type"},
            "total_quantity": {"$sum": "$quantity"},
            "count": {"$sum": 1}
        }},
    ]
    rows = await db.brand_stock_movements.aggregate(pipeline).to_list(1000)
    # Frontend dostu format
    result = {}
    for r in rows:
        b = r["_id"]["brand"]
        m = r["_id"]["machine"]
        t = r["_id"]["type"]
        key = f"{b} ({m})"
        if key not in result:
            result[key] = {"brand": b, "machine": m, "in": 0, "out": 0, "adjustment": 0}
        result[key][t] = result[key].get(t, 0) + r["total_quantity"]

    # Mevcut stok bilgisi
    current_stock = await db.brand_stock.find({}, {"_id": 0, "brand": 1, "machine": 1, "color": 1, "quantity": 1}).to_list(500)
    by_brand: dict = {}
    for s in current_stock:
        by_brand[s["brand"]] = by_brand.get(s["brand"], 0) + int(s.get("quantity", 0))

    return {
        "days": days,
        "rows": list(result.values()),
        "current_stock_by_brand": by_brand,
    }


@router.get("/brand-stock/export")
async def export_brand_stock():
    """Tüm marka stoğunu + hareketleri Excel olarak export et."""
    wb = Workbook()
    # Sheet 1: Mevcut Stok
    ws1 = wb.active
    ws1.title = "Mevcut Stok"
    headers1 = ["Marka", "Makine", "Renk", "Adet", "Notlar", "Oluşturulma", "Güncelleme"]
    ws1.append(headers1)
    for col_idx, _ in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    stocks = await db.brand_stock.find({}, {"_id": 0}).sort([("brand", 1), ("machine", 1)]).to_list(500)
    for s in stocks:
        ws1.append([
            s.get("brand", ""), s.get("machine", ""), s.get("color", "") or "",
            int(s.get("quantity", 0)), s.get("notes", "") or "",
            (s.get("created_at", "") or "")[:19].replace("T", " "),
            (s.get("updated_at", "") or "")[:19].replace("T", " "),
        ])
    for col_idx in range(1, len(headers1) + 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = 18

    # Sheet 2: Hareketler
    ws2 = wb.create_sheet("Hareketler")
    headers2 = ["Tarih", "Marka", "Makine", "Renk", "Tip", "Adet", "Müşteri", "Kullanıcı", "Not"]
    ws2.append(headers2)
    for col_idx, _ in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    movements = await db.brand_stock_movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    type_label = {"in": "Giriş", "out": "Satış", "adjustment": "Düzeltme"}
    for m in movements:
        ws2.append([
            (m.get("created_at", "") or "")[:19].replace("T", " "),
            m.get("brand", ""), m.get("machine", ""), m.get("color", "") or "",
            type_label.get(m.get("movement_type", ""), m.get("movement_type", "")),
            int(m.get("quantity", 0)),
            m.get("customer_name", "") or "",
            m.get("user_name", "") or "",
            m.get("note", "") or "",
        ])
    for col_idx in range(1, len(headers2) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"marka_stok_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
