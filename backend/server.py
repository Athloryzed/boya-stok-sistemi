from fastapi import FastAPI, APIRouter, HTTPException, Body, WebSocket, WebSocketDisconnect, UploadFile, File
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Set
import uuid
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
import json
import base64
from twilio.rest import Client as TwilioClient
import firebase_admin
from firebase_admin import credentials, messaging

# Logging'i erken yapılandır
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Firebase Admin SDK Setup
firebase_app = None
try:
    # Firebase için servis hesabı veya proje ID kullan
    firebase_config = {
        "type": "service_account",
        "project_id": "buse-kagit",
    }
    
    # Eğer servis hesabı JSON yoksa, varsayılan kimlik bilgilerini kullan
    if not firebase_admin._apps:
        # Application Default Credentials veya minimal config ile başlat
        try:
            cred = credentials.Certificate(ROOT_DIR / 'firebase-service-account.json')
            firebase_app = firebase_admin.initialize_app(cred)
            logger.info("Firebase Admin SDK initialized with service account!")
        except Exception:
            # Servis hesabı yoksa, sadece proje ID ile başlat (sınırlı özellikler)
            firebase_app = firebase_admin.initialize_app(options={'projectId': 'buse-kagit'})
            logger.info("Firebase Admin SDK initialized with project ID only")
except Exception as e:
    logger.warning(f"Firebase Admin SDK initialization failed: {e}")

async def send_fcm_notification(tokens: List[str], title: str, body: str, data: dict = None):
    """Firebase Cloud Messaging ile bildirim gönder"""
    if not tokens:
        logger.warning("No FCM tokens to send notification")
        return False
    
    try:
        message = messaging.MulticastMessage(
            notification=messaging.Notification(
                title=title,
                body=body,
            ),
            data=data or {},
            tokens=tokens,
            android=messaging.AndroidConfig(
                priority='high',
                notification=messaging.AndroidNotification(
                    sound='default',
                    priority='high',
                    channel_id='job_notifications'
                )
            ),
            webpush=messaging.WebpushConfig(
                notification=messaging.WebpushNotification(
                    icon='/logo192.png',
                    badge='/logo192.png',
                    vibrate=[200, 100, 200],
                    require_interaction=True
                )
            )
        )
        
        response = messaging.send_each_for_multicast(message)
        logger.info(f"FCM notification sent: {response.success_count} success, {response.failure_count} failed")
        return True
    except Exception as e:
        logger.error(f"FCM notification error: {e}")
        return False

# Twilio WhatsApp Setup
twilio_client = None
try:
    twilio_sid = os.environ.get('TWILIO_ACCOUNT_SID')
    twilio_token = os.environ.get('TWILIO_AUTH_TOKEN')
    logger.info(f"Twilio SID: {twilio_sid[:10] if twilio_sid else 'NOT SET'}...")
    logger.info(f"Twilio Token: {twilio_token[:10] if twilio_token else 'NOT SET'}...")
    if twilio_sid and twilio_token:
        twilio_client = TwilioClient(twilio_sid, twilio_token)
        logger.info("Twilio client initialized successfully!")
    else:
        logger.warning("Twilio credentials missing - WhatsApp disabled")
except Exception as e:
    logger.warning(f"Twilio initialization failed: {e}")

async def send_whatsapp_notification(message: str):
    """WhatsApp bildirimi gönder"""
    if not twilio_client:
        logging.warning("Twilio client not available")
        return False
    
    try:
        whatsapp_from = os.environ.get('TWILIO_WHATSAPP_FROM', 'whatsapp:+14155238886')
        whatsapp_to = os.environ.get('WHATSAPP_NOTIFY_NUMBER')
        
        if not whatsapp_to:
            logging.warning("WHATSAPP_NOTIFY_NUMBER not set")
            return False
        
        # Numara formatını düzelt
        if not whatsapp_to.startswith('whatsapp:'):
            whatsapp_to = f"whatsapp:{whatsapp_to}"
        
        logging.info(f"Sending WhatsApp to {whatsapp_to}")
        
        # Twilio sync call'ı thread'de çalıştır
        import asyncio
        loop = asyncio.get_event_loop()
        msg = await loop.run_in_executor(
            None,
            lambda: twilio_client.messages.create(
                body=message,
                from_=whatsapp_from,
                to=whatsapp_to
            )
        )
        logging.info(f"WhatsApp message sent: {msg.sid}")
        return True
    except Exception as e:
        logging.error(f"WhatsApp send failed: {e}")
        return False

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Uploads klasörü için static files
UPLOADS_DIR = Path(__file__).parent / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

# Health check endpoint for Kubernetes
@app.get("/health")
async def health_check():
    """Health check endpoint for Kubernetes liveness/readiness probes"""
    try:
        # MongoDB bağlantısını kontrol et
        await client.admin.command('ping')
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        logging.error(f"Health check failed: {e}")
        return {"status": "healthy", "database": "disconnected"}

# API health check (for /api/health route)
@api_router.get("/health")
async def api_health_check():
    """Health check endpoint via API router"""
    try:
        await client.admin.command('ping')
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        logging.error(f"Health check failed: {e}")
        return {"status": "healthy", "database": "disconnected"}

# WebSocket bağlantı yöneticisi
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []
    
    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
        logging.info(f"WebSocket connected. Total connections: {len(self.active_connections)}")
    
    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)
        logging.info(f"WebSocket disconnected. Total connections: {len(self.active_connections)}")
    
    async def broadcast(self, message: dict):
        """Tüm bağlı istemcilere mesaj gönder"""
        disconnected = []
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except Exception as e:
                logging.error(f"WebSocket send error: {e}")
                disconnected.append(connection)
        
        # Bağlantısı kopanları temizle
        for conn in disconnected:
            self.disconnect(conn)

manager = ConnectionManager()

# Yönetici WebSocket bağlantı yöneticisi
class ManagerConnectionManager:
    def __init__(self):
        self.active_connections: dict = {}  # manager_id -> WebSocket
    
    async def connect(self, websocket: WebSocket, manager_id: str):
        await websocket.accept()
        self.active_connections[manager_id] = websocket
        logging.info(f"Manager WebSocket connected: {manager_id}. Total: {len(self.active_connections)}")
    
    def disconnect(self, manager_id: str):
        if manager_id in self.active_connections:
            del self.active_connections[manager_id]
        logging.info(f"Manager WebSocket disconnected: {manager_id}. Total: {len(self.active_connections)}")
    
    async def broadcast_to_managers(self, message: dict):
        """Tüm yöneticilere mesaj gönder"""
        disconnected = []
        for manager_id, connection in self.active_connections.items():
            try:
                await connection.send_json(message)
                logging.info(f"Notification sent to manager: {manager_id}")
            except Exception as e:
                logging.error(f"Manager WebSocket send error: {e}")
                disconnected.append(manager_id)
        
        for mgr_id in disconnected:
            self.disconnect(mgr_id)

manager_ws = ManagerConnectionManager()

class Machine(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    status: str = "idle"
    current_job_id: Optional[str] = None
    maintenance: bool = False
    maintenance_reason: Optional[str] = None
    maintenance_started: Optional[str] = None

class Job(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    koli_count: int
    colors: str
    machine_id: str
    machine_name: str
    format: Optional[str] = None
    notes: Optional[str] = None
    delivery_date: Optional[str] = None
    delivery_address: Optional[str] = None
    delivery_phone: Optional[str] = None
    image_url: Optional[str] = None  # İş görseli URL'i
    status: str = "pending"  # pending, in_progress, paused, completed
    operator_name: Optional[str] = None
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    completed_koli: int = 0
    remaining_koli: int = 0  # Kalan koli (vardiya bitişinde)
    order: int = 0  # Sıra numarası (düşük = öncelikli)
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    queued_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())  # Sıraya eklenme tarihi
    # Durdurma bilgileri
    paused_at: Optional[str] = None
    pause_reason: Optional[str] = None
    produced_before_pause: int = 0  # Durdurmadan önce üretilen miktar

class Shift(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    started_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    ended_at: Optional[str] = None
    status: str = "active"  # active, pending_reports, ended
    pending_approval: bool = False  # Onay bekliyor mu?

# Vardiya Sonu Operatör Raporu (Onay Bekleyen)
class ShiftEndOperatorReport(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shift_id: str
    operator_id: str
    operator_name: str
    machine_id: str
    machine_name: str
    job_id: Optional[str] = None
    job_name: Optional[str] = None
    target_koli: int = 0
    produced_koli: int = 0
    defect_kg: float = 0.0
    is_completed: bool = False  # İş tamamlandı mı?
    status: str = "pending"  # pending, approved, rejected
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    approved_at: Optional[str] = None
    approved_by: Optional[str] = None

# Defo (Defect) Takip Modeli
class DefectLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    machine_id: str
    machine_name: str
    shift_id: Optional[str] = None
    defect_kg: float = 0.0  # Kilo cinsinden defo
    date: str = Field(default_factory=lambda: datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Vardiya Sonu Raporu
class ShiftEndReport(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shift_id: str
    machine_id: str
    machine_name: str
    job_id: Optional[str] = None
    job_name: Optional[str] = None
    target_koli: int = 0
    produced_koli: int = 0
    remaining_koli: int = 0
    defect_kg: float = 0.0  # Kilo cinsinden defo
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MaintenanceLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    machine_id: str
    machine_name: str
    reason: str
    started_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    ended_at: Optional[str] = None

class WarehouseRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    operator_name: str
    machine_name: str
    item_type: str
    quantity: int
    status: str = "pending"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PalletScan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    pallet_code: str
    job_id: str
    job_name: str
    operator_name: str
    scanned_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Boya (Paint) Models
class Paint(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    stock_kg: float = 0.0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PaintMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    paint_id: str
    paint_name: str
    movement_type: str  # "add", "remove", "to_machine", "from_machine", "used"
    amount_kg: float
    machine_id: Optional[str] = None
    machine_name: Optional[str] = None
    note: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Makineye Verilen Aktif Boya Takibi
class ActivePaintToMachine(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    paint_id: str
    paint_name: str
    machine_id: str
    machine_name: str
    given_amount_kg: float  # Verilen miktar
    returned: bool = False  # Geri alındı mı?
    returned_amount_kg: float = 0.0  # Geri alınan miktar
    used_amount_kg: float = 0.0  # Kullanılan (fark)
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    returned_at: Optional[str] = None

# Makine Mesaj Modeli
class MachineMessage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    machine_id: str
    machine_name: str
    sender_role: str  # "yonetim" veya "plan"
    sender_name: str
    message: str
    is_read: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Kullanıcı Modeli (Merkezi Yönetim)
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password: str
    role: str  # "operator", "plan", "depo", "sofor"
    display_name: Optional[str] = None
    phone: Optional[str] = None
    is_active: bool = True
    current_location_lat: Optional[float] = None
    current_location_lng: Optional[float] = None
    location_updated_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Ziyaretçi Takip Modeli
class Visitor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ip_address: str
    user_agent: str
    device_type: str  # "mobile", "tablet", "desktop"
    device_model: str
    browser: str
    os: str
    page_visited: str
    visited_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Operatör Oturum Modeli
class OperatorSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    device_id: str  # Cihaz tanımlayıcı (fingerprint)
    operator_name: str
    machine_id: Optional[str] = None
    machine_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    last_active: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    expires_at: str = Field(default_factory=lambda: (datetime.now(timezone.utc).replace(hour=23, minute=59, second=59)).isoformat())

# Palet Modeli
class Pallet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    code: str  # Palet kodu/barkod
    job_id: str
    job_name: str
    machine_id: str
    machine_name: str
    koli_count: int  # Bu paletteki koli sayısı
    operator_name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    status: str = "in_warehouse"  # in_warehouse, in_shipment, delivered

# Araç Modeli
class Vehicle(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    plate: str  # Plaka
    driver_name: Optional[str] = None
    is_active: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Sevkiyat Modeli
class Shipment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vehicle_id: str
    vehicle_plate: str
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None
    pallets: List[str] = []  # Palet ID'leri
    total_koli: int = 0
    delivery_address: str
    delivery_phone: Optional[str] = None
    delivery_notes: Optional[str] = None
    status: str = "preparing"  # preparing, in_transit, delivered, failed
    delivery_status_reason: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    created_by: str  # Oluşturan kişi (plan)

# Şoför Modeli
class Driver(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    password: str
    phone: Optional[str] = None
    is_active: bool = True
    current_location_lat: Optional[float] = None
    current_location_lng: Optional[float] = None
    location_updated_at: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Depo Sevkiyat Kaydı
class WarehouseShipmentLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    shipment_id: str
    vehicle_plate: str
    pallet_ids: List[str] = []
    total_koli: int
    partial: bool = False  # Kısmi teslim mi?
    delivered_koli: int = 0  # Teslim edilen koli sayısı
    operator_name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Audit Log Modeli
class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user: str
    action: str  # "create", "update", "delete", "start", "complete", "pause", "resume"
    entity_type: str  # "job", "user", "machine", "pallet", "shift", "paint"
    entity_name: str = ""
    details: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

async def log_audit(user: str, action: str, entity_type: str, entity_name: str = "", details: str = ""):
    """Kullanici hareket logu kaydet"""
    try:
        log = AuditLog(user=user, action=action, entity_type=entity_type, entity_name=entity_name, details=details)
        await db.audit_logs.insert_one(log.model_dump())
    except Exception as e:
        logger.error(f"Audit log error: {e}")

@api_router.get("/")
async def root():
    return {"message": "Buse Kağıt API"}

@api_router.get("/audit-logs")
async def get_audit_logs(limit: int = 100, skip: int = 0):
    """Kullanici hareket loglarini getir (en yeniden eskiye)"""
    logs = await db.audit_logs.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.audit_logs.count_documents({})
    return {"logs": logs, "total": total}

@api_router.post("/machines/init")
async def init_machines():
    machine_names = ["40x40", "40x40 ICM", "33x33 (Büyük)", "33x33 ICM", "33x33 (Eski)", "30x30", "24x24", "Dispanser"]
    existing = await db.machines.count_documents({})
    if existing == 0:
        machines = [Machine(name=name).model_dump() for name in machine_names]
        await db.machines.insert_many(machines)
    return {"message": "Machines initialized"}

@api_router.get("/machines", response_model=List[Machine])
async def get_machines():
    machines = await db.machines.find({}, {"_id": 0}).to_list(100)
    return machines

@api_router.put("/machines/{machine_id}/maintenance")
async def toggle_maintenance(machine_id: str, data: dict = Body(...)):
    maintenance = data.get("maintenance", False)
    reason = data.get("reason", "")
    
    machine = await db.machines.find_one({"id": machine_id}, {"_id": 0})
    if not machine:
        raise HTTPException(status_code=404, detail="Machine not found")
    
    update_data = {"maintenance": maintenance}
    
    if maintenance:
        update_data["maintenance_reason"] = reason
        update_data["maintenance_started"] = datetime.now(timezone.utc).isoformat()
        log = MaintenanceLog(
            machine_id=machine_id,
            machine_name=machine["name"],
            reason=reason
        )
        await db.maintenance_logs.insert_one(log.model_dump())
    else:
        update_data["maintenance_reason"] = None
        update_data["maintenance_started"] = None
        await db.maintenance_logs.update_one(
            {"machine_id": machine_id, "ended_at": None},
            {"$set": {"ended_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    await db.machines.update_one({"id": machine_id}, {"$set": update_data})
    return {"message": "Maintenance status updated"}

@api_router.get("/maintenance-logs", response_model=List[MaintenanceLog])
async def get_maintenance_logs():
    logs = await db.maintenance_logs.find({}, {"_id": 0}).sort("started_at", -1).to_list(100)
    return logs

@api_router.post("/jobs", response_model=Job)
async def create_job(job: Job):
    doc = job.model_dump()
    await db.jobs.insert_one(doc)
    
    await log_audit("Plan", "create", "job", job.name, f"Makine: {job.machine_name}, Koli: {job.koli_count}")
    
    # FCM Push Bildirimi gönder (ilgili makinedeki operatörlere)
    try:
        await send_notification_to_operators(
            machine_id=job.machine_id,
            title="📋 Yeni İş Atandı",
            body=f"{job.name} - {job.machine_name}\n📦 {job.koli_count} koli",
            data={"type": "new_job", "job_id": job.id, "machine_id": job.machine_id}
        )
    except Exception as e:
        logging.error(f"FCM notification error for new job: {e}")
    
    return job

@api_router.post("/machines/cleanup")
async def cleanup_duplicate_machines():
    machines = await db.machines.find({}, {"_id": 0}).to_list(1000)
    seen_names = {}
    duplicates_to_delete = []
    
    for machine in machines:
        if machine["name"] in seen_names:
            duplicates_to_delete.append(machine["id"])
        else:
            seen_names[machine["name"]] = machine["id"]
    
    if duplicates_to_delete:
        await db.machines.delete_many({"id": {"$in": duplicates_to_delete}})
    
    return {"message": f"Cleaned up {len(duplicates_to_delete)} duplicate machines"}

# Dosya Yükleme Endpoint'i
@api_router.post("/upload/image")
async def upload_image(file: UploadFile = File(...)):
    """Görsel yükle ve Base64 olarak MongoDB'ye kaydet"""
    try:
        # Dosya uzantısı kontrolü
        allowed_extensions = [".jpg", ".jpeg", ".png", ".gif", ".webp"]
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in allowed_extensions:
            raise HTTPException(status_code=400, detail="Sadece resim dosyaları yüklenebilir (jpg, jpeg, png, gif, webp)")
        
        # Dosya boyutu kontrolü (max 5MB)
        content = await file.read()
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Dosya boyutu 5MB'dan küçük olmalı")
        
        # Base64'e çevir
        import base64
        base64_data = base64.b64encode(content).decode('utf-8')
        
        # MIME type belirle
        mime_types = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp"
        }
        mime_type = mime_types.get(file_ext, "image/jpeg")
        
        # Data URL formatında döndür (direkt img src'de kullanılabilir)
        data_url = f"data:{mime_type};base64,{base64_data}"
        
        # MongoDB'ye kaydet (opsiyonel - gerekirse referans için)
        image_id = str(uuid.uuid4())
        await db.images.insert_one({
            "id": image_id,
            "filename": file.filename,
            "mime_type": mime_type,
            "data": data_url,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        
        return {
            "success": True,
            "filename": file.filename,
            "url": data_url,
            "image_id": image_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/jobs", response_model=List[Job])
async def get_jobs(status: Optional[str] = None, machine_id: Optional[str] = None, search: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    if machine_id:
        query["machine_id"] = machine_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"colors": {"$regex": search, "$options": "i"}}
        ]
    jobs = await db.jobs.find(query, {"_id": 0}).sort("created_at", 1).to_list(1000)
    return jobs

@api_router.post("/jobs/{job_id}/clone", response_model=Job)
async def clone_job(job_id: str, updates: dict = Body(...)):
    original_job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not original_job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    new_job = Job(
        name=updates.get("name", original_job["name"]),
        koli_count=updates.get("koli_count", original_job["koli_count"]),
        colors=updates.get("colors", original_job["colors"]),
        machine_id=updates.get("machine_id", original_job["machine_id"]),
        machine_name=updates.get("machine_name", original_job["machine_name"]),
        format=updates.get("format", original_job.get("format")),
        notes=updates.get("notes", original_job.get("notes")),
        delivery_date=updates.get("delivery_date", original_job.get("delivery_date"))
    )
    
    doc = new_job.model_dump()
    await db.jobs.insert_one(doc)
    return new_job


# Batch reorder - MUST be before /jobs/{job_id} to avoid wildcard conflict
@api_router.put("/jobs/reorder-batch")
async def reorder_jobs_batch(data: dict = Body(...)):
    """Birden fazla işin sırasını değiştir"""
    job_orders = data.get("jobs", [])
    for item in job_orders:
        await db.jobs.update_one(
            {"id": item["job_id"]},
            {"$set": {"order": item["order"]}}
        )
    return {"success": True}




@api_router.put("/jobs/{job_id}", response_model=Job)
async def update_job(job_id: str, updates: dict = Body(...)):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    await db.jobs.update_one({"id": job_id}, {"$set": updates})
    
    await log_audit("Yonetim", "update", "job", job.get("name", ""), f"Guncellenen: {', '.join(updates.keys())}")
    
    updated_job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    return Job(**updated_job)

@api_router.delete("/jobs/{job_id}")
async def delete_job(job_id: str):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    result = await db.jobs.delete_one({"id": job_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")
    await log_audit("Yonetim", "delete", "job", job.get("name", "") if job else job_id)
    return {"message": "Job deleted"}

@api_router.put("/jobs/{job_id}/start")
async def start_job(job_id: str, data: dict = Body(...)):
    operator_name = data.get("operator_name")
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {
            "status": "in_progress",
            "operator_name": operator_name,
            "started_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await db.machines.update_one(
        {"id": job["machine_id"]},
        {"$set": {"status": "working", "current_job_id": job_id}}
    )
    
    await log_audit(operator_name or "Operator", "start", "job", job.get("name", ""), f"Makine: {job.get('machine_name', '')}")
    
    return {"message": "Job started"}

@api_router.put("/jobs/{job_id}/complete")
async def complete_job(job_id: str, data: dict = Body(None)):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    completed_koli = data.get("completed_koli", job["koli_count"]) if data else job["koli_count"]
    
    # Önce veritabanını güncelle - bu kritik işlem
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "completed_koli": completed_koli
        }}
    )
    
    await db.machines.update_one(
        {"id": job["machine_id"]},
        {"$set": {"status": "idle", "current_job_id": None}}
    )
    
    await log_audit(job.get("operator_name", "Operator"), "complete", "job", job.get("name", ""), f"Koli: {completed_koli}")
    
    # Hızlı response dön - bildirimleri arka planda gönder
    # Background task olarak bildirim gönder
    asyncio.create_task(send_completion_notifications(job, job_id, completed_koli))
    
    return {"success": True, "message": "İş tamamlandı"}

async def send_completion_notifications(job: dict, job_id: str, completed_koli: int):
    """Bildirimler arka planda gönderilir - UI'ı bloklamaz"""
    try:
        notification_title = "✅ İş Tamamlandı!"
        notification_body = f"📋 {job['name']}\n🏭 {job['machine_name']}\n📦 {completed_koli} koli"
        message = f"✅ İş Tamamlandı!\n\n📋 İş: {job['name']}\n🏭 Makine: {job['machine_name']}\n📦 Koli: {completed_koli}\n👷 Operatör: {job.get('operator_name', '-')}\n⏰ Tarih: {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')}"
        
        # FCM bildirimleri paralel gönder
        await asyncio.gather(
            send_notification_to_managers(
                title=notification_title,
                body=notification_body,
                data={"type": "job_completed", "job_id": job_id}
            ),
            send_notification_to_plan_users(
                title=notification_title,
                body=notification_body,
                data={"type": "job_completed", "job_id": job_id}
            ),
            return_exceptions=True
        )
        
        # WebSocket bildirimi
        await manager_ws.broadcast_to_managers({
            "type": "job_completed",
            "message": message,
            "job_name": job['name'],
            "machine_name": job['machine_name'],
            "completed_koli": completed_koli
        })
    except Exception as e:
        logging.error(f"Background notification error: {e}")

# İş Durdurma
@api_router.put("/jobs/{job_id}/pause")
async def pause_job(job_id: str, data: dict = Body(...)):
    """İşi durdur ve sebep not et"""
    pause_reason = data.get("pause_reason", "")
    produced_koli = data.get("produced_koli", 0)  # Durdurmadan önce üretilen
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="İş bulunamadı")
    
    if job["status"] != "in_progress":
        raise HTTPException(status_code=400, detail="Sadece devam eden işler durdurulabilir")
    
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {
            "status": "paused",
            "paused_at": datetime.now(timezone.utc).isoformat(),
            "pause_reason": pause_reason,
            "produced_before_pause": produced_koli
        }}
    )
    
    # Makineyi boşalt
    await db.machines.update_one(
        {"id": job["machine_id"]},
        {"$set": {"status": "idle", "current_job_id": None}}
    )
    
    # WebSocket bildirimi
    await manager.broadcast({
        "type": "job_paused",
        "data": {
            "job_id": job_id,
            "job_name": job["name"],
            "machine_id": job["machine_id"],
            "pause_reason": pause_reason
        }
    })
    
    return {"message": "İş durduruldu", "job_id": job_id}
    await log_audit(job.get("operator_name", "Operator"), "pause", "job", job.get("name", ""), f"Sebep: {pause_reason}")


# Durdurulan İşe Devam Et
@api_router.put("/jobs/{job_id}/resume")
async def resume_job(job_id: str, data: dict = Body(...)):
    """Durdurulan işe devam et"""
    operator_name = data.get("operator_name", "")
    
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="İş bulunamadı")
    
    if job["status"] != "paused":
        raise HTTPException(status_code=400, detail="Sadece durdurulmuş işlere devam edilebilir")
    
    # Makinede başka aktif iş var mı kontrol et
    active_on_machine = await db.jobs.find_one({
        "machine_id": job["machine_id"],
        "status": "in_progress"
    })
    if active_on_machine:
        raise HTTPException(status_code=400, detail="Bu makinede zaten aktif bir iş var")
    
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {
            "status": "in_progress",
            "operator_name": operator_name or job.get("operator_name"),
            "started_at": datetime.now(timezone.utc).isoformat()  # Yeni başlangıç zamanı
        }}
    )
    
    await db.machines.update_one(
        {"id": job["machine_id"]},
        {"$set": {"status": "working", "current_job_id": job_id}}
    )
    
    await log_audit(operator_name or "Operator", "resume", "job", job.get("name", ""), f"Makine: {job.get('machine_name', '')}")

    return {"message": "İşe devam edildi", "job_id": job_id}

# Durdurulan İşleri Listele
@api_router.get("/jobs/paused")
async def get_paused_jobs():
    """Durdurulmuş işleri listele"""
    paused = await db.jobs.find({"status": "paused"}, {"_id": 0}).to_list(100)
    return paused

# İş Sırası Değiştirme
@api_router.put("/jobs/{job_id}/reorder")
async def reorder_job(job_id: str, data: dict = Body(...)):
    """İşin sırasını değiştir"""
    new_order = data.get("order", 0)
    
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {"order": new_order}}
    )
    return {"success": True}

# ==================== YENİ VARDİYA SONU AKIŞI ====================

@api_router.post("/shifts/request-end")
async def request_shift_end():
    """Vardiya sonu bildirimi gönder - tüm aktif operatörlere"""
    active_shift = await db.shifts.find_one({"status": "active"}, {"_id": 0})
    if not active_shift:
        raise HTTPException(status_code=400, detail="Aktif vardiya bulunamadı")
    
    # Vardiyayı "pending_reports" durumuna al
    await db.shifts.update_one(
        {"id": active_shift["id"]},
        {"$set": {"status": "pending_reports", "pending_approval": True}}
    )
    
    # Aktif işleri olan makineleri bul
    active_jobs = await db.jobs.find({"status": "in_progress"}, {"_id": 0}).to_list(100)
    
    # Her aktif iş için operatöre bildirim gönder (WebSocket üzerinden)
    notifications_sent = []
    for job in active_jobs:
        notification = {
            "type": "shift_end_request",
            "shift_id": active_shift["id"],
            "job_id": job["id"],
            "job_name": job["name"],
            "machine_id": job["machine_id"],
            "machine_name": job["machine_name"],
            "target_koli": job["koli_count"],
            "operator_name": job.get("operator_name", ""),
            "message": "Vardiya bitti! Lütfen işinizi tamamlayın veya üretim bilgilerinizi girin."
        }
        notifications_sent.append(notification)
        
        # WebSocket üzerinden bildirim gönder
        await manager.broadcast({
            "type": "shift_end_request",
            "data": notification
        })
    
    return {
        "message": "Vardiya sonu bildirimi gönderildi",
        "shift_id": active_shift["id"],
        "notifications_sent": len(notifications_sent),
        "active_jobs": notifications_sent
    }

@api_router.post("/shifts/operator-report")
async def submit_operator_report(data: dict = Body(...)):
    """Operatörün vardiya sonu raporu - onay için gönder"""
    shift_id = data.get("shift_id")
    operator_id = data.get("operator_id", "")
    operator_name = data.get("operator_name", "")
    machine_id = data.get("machine_id")
    machine_name = data.get("machine_name", "")
    job_id = data.get("job_id")
    job_name = data.get("job_name", "")
    target_koli = data.get("target_koli", 0)
    produced_koli = data.get("produced_koli", 0)
    defect_kg = float(data.get("defect_kg", 0))
    is_completed = data.get("is_completed", False)
    
    # Rapor oluştur
    report = ShiftEndOperatorReport(
        shift_id=shift_id,
        operator_id=operator_id,
        operator_name=operator_name,
        machine_id=machine_id,
        machine_name=machine_name,
        job_id=job_id,
        job_name=job_name,
        target_koli=target_koli,
        produced_koli=produced_koli,
        defect_kg=defect_kg,
        is_completed=is_completed,
        status="pending"
    )
    
    await db.shift_operator_reports.insert_one(report.model_dump())
    
    # WebSocket ile yönetime bildir
    await manager.broadcast({
        "type": "new_operator_report",
        "data": {
            "report_id": report.id,
            "operator_name": operator_name,
            "machine_name": machine_name,
            "job_name": job_name
        }
    })
    
    return {"message": "Rapor gönderildi, onay bekleniyor", "report_id": report.id}

@api_router.get("/shifts/pending-reports")
async def get_pending_reports():
    """Onay bekleyen operatör raporlarını listele"""
    reports = await db.shift_operator_reports.find(
        {"status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return reports

@api_router.post("/shifts/approve-report/{report_id}")
async def approve_operator_report(report_id: str, data: dict = Body(None)):
    """Operatör raporunu onayla"""
    approved_by = data.get("approved_by", "Yönetim") if data else "Yönetim"
    
    report = await db.shift_operator_reports.find_one({"id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Rapor bulunamadı")
    
    # Raporu onayla
    await db.shift_operator_reports.update_one(
        {"id": report_id},
        {"$set": {
            "status": "approved",
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "approved_by": approved_by
        }}
    )
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # İş tamamlandıysa
    if report.get("is_completed"):
        job_id = report.get("job_id")
        if job_id:
            job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
            if job:
                completed_koli = report.get("target_koli", job.get("koli_count", 0))
                await db.jobs.update_one(
                    {"id": job_id},
                    {"$set": {
                        "status": "completed",
                        "completed_at": datetime.now(timezone.utc).isoformat(),
                        "completed_koli": completed_koli
                    }}
                )
                await db.machines.update_one(
                    {"id": report["machine_id"]},
                    {"$set": {"status": "idle", "current_job_id": None}}
                )
                
                # WhatsApp bildirimi
                try:
                    message = f"✅ İş Tamamlandı!\n\n📋 İş: {job['name']}\n🏭 Makine: {job['machine_name']}\n📦 Koli: {completed_koli}\n👷 Operatör: {report.get('operator_name', '-')}"
                    await send_whatsapp_notification(message)
                except Exception as e:
                    logging.error(f"WhatsApp error: {e}")
    else:
        # Kısmi üretim - shift_end_reports'a kaydet
        shift_report = ShiftEndReport(
            shift_id=report["shift_id"],
            machine_id=report["machine_id"],
            machine_name=report["machine_name"],
            job_id=report.get("job_id"),
            job_name=report.get("job_name"),
            target_koli=report.get("target_koli", 0),
            produced_koli=report.get("produced_koli", 0),
            remaining_koli=report.get("target_koli", 0) - report.get("produced_koli", 0),
            defect_kg=report.get("defect_kg", 0)
        )
        await db.shift_end_reports.insert_one(shift_report.model_dump())
        
        # İşin kalan kolisini güncelle
        if report.get("job_id"):
            remaining = report.get("target_koli", 0) - report.get("produced_koli", 0)
            await db.jobs.update_one(
                {"id": report["job_id"]},
                {"$set": {
                    "remaining_koli": remaining if remaining > 0 else 0,
                    "completed_koli": report.get("produced_koli", 0),
                    "status": "pending"  # Tekrar beklemede
                }}
            )
        
        # Defo kaydı
        if report.get("defect_kg", 0) > 0:
            defect_log = DefectLog(
                machine_id=report["machine_id"],
                machine_name=report["machine_name"],
                shift_id=report["shift_id"],
                defect_kg=report["defect_kg"],
                date=today
            )
            await db.defect_logs.insert_one(defect_log.model_dump())
        
        # Makineyi idle yap
        await db.machines.update_one(
            {"id": report["machine_id"]},
            {"$set": {"status": "idle", "current_job_id": None}}
        )
    
    return {"message": "Rapor onaylandı"}

@api_router.post("/shifts/approve-all")
async def approve_all_reports_and_end_shift():
    """Tüm raporları onayla ve vardiyayı bitir"""
    pending_reports = await db.shift_operator_reports.find(
        {"status": "pending"},
        {"_id": 0}
    ).to_list(100)
    
    for report in pending_reports:
        await approve_operator_report(report["id"], {"approved_by": "Yönetim (Toplu)"})
    
    # Vardiyayı bitir
    active_shift = await db.shifts.find_one({"status": "pending_reports"}, {"_id": 0})
    if active_shift:
        await db.shifts.update_one(
            {"id": active_shift["id"]},
            {"$set": {
                "status": "ended",
                "ended_at": datetime.now(timezone.utc).isoformat(),
                "pending_approval": False
            }}
        )
    
    return {"message": f"{len(pending_reports)} rapor onaylandı ve vardiya bitirildi"}

@api_router.get("/shifts/status")
async def get_shift_status():
    """Mevcut vardiya durumunu getir"""
    shift = await db.shifts.find_one(
        {"status": {"$in": ["active", "pending_reports"]}},
        {"_id": 0}
    )
    if not shift:
        return {"status": "no_active_shift", "shift": None}
    
    pending_count = await db.shift_operator_reports.count_documents({"status": "pending", "shift_id": shift["id"]})
    
    return {
        "status": shift["status"],
        "shift": shift,
        "pending_reports_count": pending_count
    }

# Vardiya Sonu Raporu (Eski - Geriye Uyumluluk)
@api_router.post("/shifts/end-with-report")
async def end_shift_with_report(data: dict = Body(...)):
    """Vardiya bitişinde makine bazlı üretim ve defo raporu"""
    machine_reports = data.get("reports", [])
    # Her rapor: {machine_id, machine_name, job_id, job_name, target_koli, produced_koli, defect_kg}
    
    active_shift = await db.shifts.find_one({"status": "active"}, {"_id": 0})
    if not active_shift:
        raise HTTPException(status_code=400, detail="Aktif vardiya bulunamadı")
    
    shift_id = active_shift["id"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    for report in machine_reports:
        machine_id = report.get("machine_id")
        machine_name = report.get("machine_name", "")
        job_id = report.get("job_id")
        job_name = report.get("job_name", "")
        target_koli = report.get("target_koli", 0)
        produced_koli = report.get("produced_koli", 0)
        defect_kg = float(report.get("defect_kg", 0))
        remaining_koli = target_koli - produced_koli
        
        # Vardiya sonu raporu kaydet
        shift_report = ShiftEndReport(
            shift_id=shift_id,
            machine_id=machine_id,
            machine_name=machine_name,
            job_id=job_id,
            job_name=job_name,
            target_koli=target_koli,
            produced_koli=produced_koli,
            remaining_koli=remaining_koli if remaining_koli > 0 else 0,
            defect_kg=defect_kg
        )
        await db.shift_end_reports.insert_one(shift_report.model_dump())
        
        # İşin kalan kolisini güncelle
        if job_id and remaining_koli > 0:
            await db.jobs.update_one(
                {"id": job_id},
                {"$set": {
                    "remaining_koli": remaining_koli,
                    "completed_koli": produced_koli
                }}
            )
        
        # Defo kaydı oluştur (kg cinsinden)
        if defect_kg > 0:
            defect_log = DefectLog(
                machine_id=machine_id,
                machine_name=machine_name,
                shift_id=shift_id,
                defect_kg=defect_kg,
                date=today
            )
            await db.defect_logs.insert_one(defect_log.model_dump())
        
        # Makineyi idle yap
        await db.machines.update_one(
            {"id": machine_id},
            {"$set": {"status": "idle", "current_job_id": None}}
        )
    
    # Vardiyayı bitir
    await db.shifts.update_one(
        {"id": shift_id},
        {"$set": {
            "status": "ended",
            "ended_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Vardiya raporu kaydedildi ve vardiya bitirildi"}

# Defo Takibi API'leri
@api_router.post("/defects")
async def create_defect_log(data: dict = Body(...)):
    """Defo kaydı oluştur"""
    defect_log = DefectLog(
        machine_id=data.get("machine_id"),
        machine_name=data.get("machine_name"),
        shift_id=data.get("shift_id"),
        defect_kg=float(data.get("defect_kg", 0)),
        notes=data.get("notes")
    )
    await db.defect_logs.insert_one(defect_log.model_dump())
    return defect_log

@api_router.get("/defects")
async def get_defect_logs(machine_id: Optional[str] = None, date: Optional[str] = None, limit: int = 100):
    """Defo kayıtlarını listele"""
    query = {}
    if machine_id:
        query["machine_id"] = machine_id
    if date:
        query["date"] = date
    
    defects = await db.defect_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return defects

@api_router.get("/defects/analytics/weekly")
async def get_defect_analytics_weekly():
    """Haftalık defo analitikleri"""
    from datetime import timedelta
    
    start_date = datetime.now(timezone.utc) - timedelta(days=7)
    start_date_str = start_date.strftime("%Y-%m-%d")
    
    defects = await db.defect_logs.find(
        {"date": {"$gte": start_date_str}},
        {"_id": 0}
    ).to_list(1000)
    
    # Makine bazlı defo istatistikleri
    machine_defects = {}
    daily_defects = {}
    total_defects = 0.0
    
    for defect in defects:
        machine = defect.get("machine_name", "Bilinmiyor")
        kg = float(defect.get("defect_kg", 0))
        date = defect.get("date", "")
        
        total_defects += kg
        
        if machine not in machine_defects:
            machine_defects[machine] = 0.0
        machine_defects[machine] += kg
        
        if date not in daily_defects:
            daily_defects[date] = 0.0
        daily_defects[date] += kg
    
    return {
        "total_defects_kg": round(total_defects, 2),
        "machine_defects": {k: round(v, 2) for k, v in machine_defects.items()},
        "daily_defects": {k: round(v, 2) for k, v in daily_defects.items()},
        "period": "weekly"
    }

@api_router.get("/defects/analytics/monthly")
async def get_defect_analytics_monthly(year: int = None, month: int = None):
    """Aylık defo analitikleri"""
    from datetime import timedelta
    import calendar
    
    if year is None:
        year = datetime.now(timezone.utc).year
    if month is None:
        month = datetime.now(timezone.utc).month
    
    # Ayın ilk ve son günü
    first_day = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day_num = calendar.monthrange(year, month)[1]
    last_day = datetime(year, month, last_day_num, 23, 59, 59, tzinfo=timezone.utc)
    
    defects = await db.defect_logs.find(
        {"date": {"$gte": first_day.strftime("%Y-%m-%d"), "$lte": last_day.strftime("%Y-%m-%d")}},
        {"_id": 0}
    ).to_list(1000)
    
    machine_defects = {}
    daily_defects = {}
    total_defects = 0.0
    
    for defect in defects:
        machine = defect.get("machine_name", "Bilinmiyor")
        kg = float(defect.get("defect_kg", 0))
        date = defect.get("date", "")
        
        total_defects += kg
        
        if machine not in machine_defects:
            machine_defects[machine] = 0.0
        machine_defects[machine] += kg
        
        if date not in daily_defects:
            daily_defects[date] = 0.0
        daily_defects[date] += kg
    
    month_names = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", 
                   "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
    
    return {
        "total_defects_kg": round(total_defects, 2),
        "machine_defects": {k: round(v, 2) for k, v in machine_defects.items()},
        "daily_defects": {k: round(v, 2) for k, v in daily_defects.items()},
        "period": "monthly",
        "month_name": month_names[month],
        "year": year,
        "month": month
    }

@api_router.get("/defects/analytics/daily-by-week")
async def get_defect_analytics_daily_by_week(week_offset: int = 0):
    """Hafta bazında günlük defo analitikleri"""
    from datetime import timedelta
    
    # Hedef haftayı hesapla
    today = datetime.now(timezone.utc)
    days_since_monday = today.weekday()
    this_monday = today - timedelta(days=days_since_monday)
    target_monday = this_monday + timedelta(weeks=week_offset)
    target_sunday = target_monday + timedelta(days=6)
    
    daily_stats = []
    day_names = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
    
    for i in range(7):
        date = target_monday + timedelta(days=i)
        date_str = date.strftime("%Y-%m-%d")
        
        defects = await db.defect_logs.find(
            {"date": date_str},
            {"_id": 0}
        ).to_list(100)
        
        total_kg = sum(float(d.get("defect_kg", 0)) for d in defects)
        machine_breakdown = {}
        for d in defects:
            machine = d.get("machine_name", "Bilinmiyor")
            if machine not in machine_breakdown:
                machine_breakdown[machine] = 0.0
            machine_breakdown[machine] += float(d.get("defect_kg", 0))
        
        daily_stats.append({
            "date": date.strftime("%d %b"),
            "day_name": day_names[i],
            "full_date": date_str,
            "total_kg": round(total_kg, 2),
            "machines": {k: round(v, 2) for k, v in machine_breakdown.items()}
        })
    
    return {
        "week_start": target_monday.strftime("%d %b %Y"),
        "week_end": target_sunday.strftime("%d %b %Y"),
        "week_offset": week_offset,
        "daily_stats": daily_stats
    }

@api_router.get("/defects/analytics")
async def get_defect_analytics(period: str = "weekly"):
    """Defo analitikleri (geriye uyumluluk)"""
    if period == "monthly":
        return await get_defect_analytics_monthly()
    return await get_defect_analytics_weekly()

@api_router.get("/shift-reports")
async def get_shift_reports(shift_id: Optional[str] = None, limit: int = 50):
    """Vardiya sonu raporlarını listele"""
    query = {}
    if shift_id:
        query["shift_id"] = shift_id
    
    reports = await db.shift_end_reports.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return reports

@api_router.post("/shifts/start")
async def start_shift():
    active_shift = await db.shifts.find_one({"status": "active"}, {"_id": 0})
    if active_shift:
        raise HTTPException(status_code=400, detail="There is already an active shift")
    
    shift = Shift()
    await db.shifts.insert_one(shift.model_dump())
    
    # Tüm çalışanlara vardiya başladı bildirimi gönder
    try:
        await send_notification_to_all_workers(
            title="🏭 Vardiya Başladı!",
            body="Günlük vardiya başlamıştır. İyi çalışmalar!",
            data={"type": "shift_started", "shift_id": shift.id}
        )
    except Exception as e:
        logging.error(f"Shift start notification error: {e}")
    
    await log_audit("Yonetim", "create", "shift", shift.id, "Vardiya baslatildi")

    return shift

@api_router.post("/shifts/end")
async def end_shift():
    active_shift = await db.shifts.find_one({"status": "active"}, {"_id": 0})
    if not active_shift:
        raise HTTPException(status_code=400, detail="No active shift found")
    
    await db.shifts.update_one(
        {"id": active_shift["id"]},
        {"$set": {
            "status": "ended",
            "ended_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Shift ended"}

# Vardiya bitirme bildirimi - operatörlere rapor doldurmaları için
@api_router.post("/shifts/notify-end")
async def notify_shift_end():
    """Vardiya bitiş bildirimi gönder - operatörler rapor dolduracak"""
    active_shift = await db.shifts.find_one({"status": "active"}, {"_id": 0})
    if not active_shift:
        raise HTTPException(status_code=400, detail="Aktif vardiya bulunamadı")
    
    # Aktif işleri al
    active_jobs = await db.jobs.find(
        {"status": {"$in": ["in_progress", "paused"]}},
        {"_id": 0}
    ).to_list(100)
    
    # Tüm operatörlere bildirim gönder
    try:
        await send_notification_to_operators(
            machine_id="all",
            title="⏰ Vardiya Bitti!",
            body="Lütfen üretim ve defo bilgilerinizi girin.",
            data={"type": "shift_end_report", "shift_id": active_shift["id"]}
        )
    except Exception as e:
        logging.error(f"Shift end notification error: {e}")
    
    return {
        "message": "Operatörlere bildirim gönderildi",
        "active_jobs": active_jobs,
        "shift_id": active_shift["id"]
    }

@api_router.get("/shifts/current")
async def get_current_shift():
    shift = await db.shifts.find_one({"status": "active"}, {"_id": 0})
    return shift

@api_router.get("/analytics/weekly")
async def get_weekly_analytics():
    from datetime import timedelta
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    week_ago_str = week_ago.isoformat()
    
    # Tamamlanan işlerden üretim
    completed_jobs = await db.jobs.find(
        {"status": "completed", "completed_at": {"$gte": week_ago_str}},
        {"_id": 0}
    ).to_list(1000)
    
    # Vardiya raporlarından kısmi üretim
    shift_reports = await db.shift_end_reports.find(
        {"created_at": {"$gte": week_ago_str}},
        {"_id": 0}
    ).to_list(1000)
    
    machine_stats = {}
    
    # Tamamlanan işler
    for job in completed_jobs:
        machine = job["machine_name"]
        koli = job.get("completed_koli", job.get("koli_count", 0))
        if machine not in machine_stats:
            machine_stats[machine] = 0
        machine_stats[machine] += koli
    
    # Vardiya raporlarından ekleme (sadece tamamlanmamış işler için) - Batch query
    job_ids = [r["job_id"] for r in shift_reports if r.get("job_id")]
    if job_ids:
        jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
        jobs_dict = {j["id"]: j for j in jobs_list}
        
        for report in shift_reports:
            job_id = report.get("job_id")
            if job_id:
                job = jobs_dict.get(job_id)
                if job and job.get("status") != "completed":
                    machine = report.get("machine_name", "")
                    koli = report.get("produced_koli", 0)
                    if machine and koli > 0:
                        if machine not in machine_stats:
                            machine_stats[machine] = 0
                        machine_stats[machine] += koli
    
    return {
        "machine_stats": machine_stats
    }

@api_router.get("/analytics/daily")
async def get_daily_analytics():
    from datetime import timedelta
    
    # Son 7 günün her günü için
    daily_stats = []
    for i in range(7):
        date = datetime.now(timezone.utc) - timedelta(days=i)
        start_of_day = date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        
        # Tamamlanan işler
        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)
        
        # Vardiya raporları
        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)
        
        total_koli = sum(job.get("completed_koli", job.get("koli_count", 0)) for job in jobs)
        machine_breakdown = {}
        
        for job in jobs:
            machine = job["machine_name"]
            if machine not in machine_breakdown:
                machine_breakdown[machine] = 0
            machine_breakdown[machine] += job.get("completed_koli", job.get("koli_count", 0))
        
        # Vardiya raporlarından kısmi üretim (tamamlanmamış işler için) - Batch query
        job_ids = [r["job_id"] for r in shift_reports if r.get("job_id")]
        if job_ids:
            jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
            jobs_dict = {j["id"]: j for j in jobs_list}
            
            for report in shift_reports:
                job_id = report.get("job_id")
                if job_id:
                    job = jobs_dict.get(job_id)
                    if job and job.get("status") != "completed":
                        machine = report.get("machine_name", "")
                        koli = report.get("produced_koli", 0)
                        if machine and koli > 0:
                            if machine not in machine_breakdown:
                                machine_breakdown[machine] = 0
                            machine_breakdown[machine] += koli
                            total_koli += koli
        
        daily_stats.append({
            "date": start_of_day.strftime("%d %b"),
            "total_koli": total_koli,
            "machines": machine_breakdown
        })
    
    return {"daily_stats": list(reversed(daily_stats))}

@api_router.get("/analytics/monthly")
async def get_monthly_analytics(year: Optional[int] = None, month: Optional[int] = None):
    from datetime import timedelta
    
    if year and month:
        from datetime import date
        if month == 12:
            start_date = datetime(year, month, 1, tzinfo=timezone.utc)
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            start_date = datetime(year, month, 1, tzinfo=timezone.utc)
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        start_date_str = start_date.isoformat()
        end_date_str = end_date.isoformat()
        
        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": start_date_str, "$lt": end_date_str}},
            {"_id": 0}
        ).to_list(1000)
        
        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": start_date_str, "$lt": end_date_str}},
            {"_id": 0}
        ).to_list(1000)
    else:
        month_ago = datetime.now(timezone.utc) - timedelta(days=30)
        month_ago_str = month_ago.isoformat()
        
        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": month_ago_str}},
            {"_id": 0}
        ).to_list(1000)
        
        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": month_ago_str}},
            {"_id": 0}
        ).to_list(1000)
    
    machine_stats = {}
    
    for job in jobs:
        machine = job["machine_name"]
        koli = job.get("completed_koli", job.get("koli_count", 0))
        
        if machine not in machine_stats:
            machine_stats[machine] = 0
        machine_stats[machine] += koli
    
    # Vardiya raporlarından kısmi üretim - Batch query
    job_ids = [r["job_id"] for r in shift_reports if r.get("job_id")]
    if job_ids:
        jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
        jobs_dict = {j["id"]: j for j in jobs_list}
        
        for report in shift_reports:
            job_id = report.get("job_id")
            if job_id:
                job = jobs_dict.get(job_id)
                if job and job.get("status") != "completed":
                    machine = report.get("machine_name", "")
                    koli = report.get("produced_koli", 0)
                    if machine and koli > 0:
                        if machine not in machine_stats:
                            machine_stats[machine] = 0
                        machine_stats[machine] += koli
    
    return {
        "machine_stats": machine_stats
    }

@api_router.get("/analytics/export")
async def export_analytics(period: str = "weekly"):
    from datetime import timedelta
    
    if period == "weekly":
        date_ago = datetime.now(timezone.utc) - timedelta(days=7)
    else:
        date_ago = datetime.now(timezone.utc) - timedelta(days=30)
    
    date_ago_str = date_ago.isoformat()
    
    jobs = await db.jobs.find(
        {"status": "completed", "completed_at": {"$gte": date_ago_str}},
        {"_id": 0}
    ).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Üretim Raporu"
    
    header_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    header_font = Font(bold=True, size=12)
    
    headers = ["İş Adı", "Makine", "Operatör", "Koli Sayısı", "Başlangıç", "Tamamlanma"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row, job in enumerate(jobs, start=2):
        ws.cell(row=row, column=1, value=job["name"])
        ws.cell(row=row, column=2, value=job["machine_name"])
        ws.cell(row=row, column=3, value=job.get("operator_name", "Unknown"))
        ws.cell(row=row, column=4, value=job["completed_koli"])
        ws.cell(row=row, column=5, value=job.get("started_at", ""))
        ws.cell(row=row, column=6, value=job.get("completed_at", ""))
    
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"uretim_raporu_{period}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/warehouse-requests", response_model=WarehouseRequest)
async def create_warehouse_request(request: WarehouseRequest):
    doc = request.model_dump()
    await db.warehouse_requests.insert_one(doc)
    
    # WebSocket ile depo'ya bildirim gönder
    await manager.broadcast({
        "type": "new_warehouse_request",
        "data": {
            "id": request.id,
            "operator_name": request.operator_name,
            "machine_name": request.machine_name,
            "item_type": request.item_type,
            "quantity": request.quantity,
            "created_at": request.created_at
        }
    })
    
    return request

@api_router.get("/warehouse-requests", response_model=List[WarehouseRequest])
async def get_warehouse_requests(status: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    requests = await db.warehouse_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return requests

@api_router.put("/warehouse-requests/{request_id}/complete")
async def complete_warehouse_request(request_id: str):
    await db.warehouse_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "completed"}}
    )
    return {"message": "Request completed"}

# scan_pallet fonksiyonu aşağıdaki create_pallet ile birleştirildi

# ==================== BOYA (PAINT) ENDPOINTS ====================

# Başlangıç boyaları (Güncellenmiş liste)
INITIAL_PAINTS = [
    "Siyah", "Beyaz", "Mavi", "Lacivert", "Refleks", "Kırmızı",
    "Magenta", "Rhodam", "Sarı", "Gold", "Gümüş", "Pasta"
]

# Düşük stok eşiği (kg/L)
LOW_STOCK_THRESHOLD = 5.0

@api_router.post("/paints/init")
async def init_paints():
    """Başlangıç boyalarını oluştur"""
    existing = await db.paints.count_documents({})
    if existing == 0:
        paints = [Paint(name=name).model_dump() for name in INITIAL_PAINTS]
        await db.paints.insert_many(paints)
        return {"message": f"{len(INITIAL_PAINTS)} boya eklendi"}
    return {"message": "Boyalar zaten mevcut"}

@api_router.get("/paints", response_model=List[Paint])
async def get_paints():
    """Tüm boyaları listele"""
    paints = await db.paints.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return paints

@api_router.post("/paints", response_model=Paint)
async def create_paint(paint: Paint):
    """Yeni boya ekle"""
    doc = paint.model_dump()
    await db.paints.insert_one(doc)
    return paint

@api_router.delete("/paints/{paint_id}")
async def delete_paint(paint_id: str):
    """Boya sil"""
    result = await db.paints.delete_one({"id": paint_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Boya bulunamadı")
    return {"message": "Boya silindi"}

@api_router.delete("/paints/movements/clear")
async def clear_paint_movements():
    """Tüm boya hareketlerini temizle"""
    result = await db.paint_movements.delete_many({})
    return {"message": f"{result.deleted_count} hareket silindi"}

@api_router.post("/paints/transaction")
async def paint_transaction(data: dict = Body(...)):
    """Boya hareketi kaydet (stok ekleme, çıkarma, makineye gönderme, makineden alma)"""
    paint_id = data.get("paint_id")
    movement_type = data.get("movement_type")  # "add", "remove", "to_machine", "from_machine"
    amount_kg = float(data.get("amount_kg", 0))
    machine_id = data.get("machine_id")
    machine_name = data.get("machine_name")
    note = data.get("note", "")
    
    # Boya bul
    paint = await db.paints.find_one({"id": paint_id}, {"_id": 0})
    if not paint:
        raise HTTPException(status_code=404, detail="Boya bulunamadı")
    
    # Stok güncelle
    current_stock = paint.get("stock_kg", 0)
    
    if movement_type == "add" or movement_type == "from_machine":
        new_stock = current_stock + amount_kg
    elif movement_type == "remove" or movement_type == "to_machine":
        if current_stock < amount_kg:
            raise HTTPException(status_code=400, detail=f"Yetersiz stok! Mevcut: {current_stock} kg")
        new_stock = current_stock - amount_kg
    else:
        raise HTTPException(status_code=400, detail="Geçersiz hareket tipi")
    
    # Stok güncelle
    await db.paints.update_one({"id": paint_id}, {"$set": {"stock_kg": new_stock}})
    
    # Hareket kaydı oluştur
    movement = PaintMovement(
        paint_id=paint_id,
        paint_name=paint["name"],
        movement_type=movement_type,
        amount_kg=amount_kg,
        machine_id=machine_id,
        machine_name=machine_name,
        note=note
    )
    await db.paint_movements.insert_one(movement.model_dump())
    
    return {
        "message": "Hareket kaydedildi",
        "new_stock": new_stock,
        "movement_id": movement.id
    }

@api_router.get("/paints/movements", response_model=List[PaintMovement])
async def get_paint_movements(paint_id: Optional[str] = None, limit: int = 100):
    """Boya hareketlerini listele"""
    query = {}
    if paint_id:
        query["paint_id"] = paint_id
    movements = await db.paint_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return movements

# Yeni Boya Takip Sistemi - Makineye Ver/Geri Al

@api_router.post("/paints/give-to-machine")
async def give_paint_to_machine(data: dict = Body(...)):
    """Makineye boya ver - tartıdan okunan değeri gir"""
    paint_id = data.get("paint_id")
    machine_id = data.get("machine_id")
    machine_name = data.get("machine_name")
    given_amount_kg = float(data.get("amount_kg", 0))
    
    if given_amount_kg <= 0:
        raise HTTPException(status_code=400, detail="Geçerli bir miktar girin")
    
    # Boya kontrol
    paint = await db.paints.find_one({"id": paint_id}, {"_id": 0})
    if not paint:
        raise HTTPException(status_code=404, detail="Boya bulunamadı")
    
    # Stok kontrol
    current_stock = paint.get("stock_kg", 0)
    if current_stock < given_amount_kg:
        raise HTTPException(status_code=400, detail=f"Yetersiz stok! Mevcut: {current_stock} kg")
    
    # Aktif boya kaydı oluştur
    active_paint = ActivePaintToMachine(
        paint_id=paint_id,
        paint_name=paint["name"],
        machine_id=machine_id,
        machine_name=machine_name,
        given_amount_kg=given_amount_kg
    )
    await db.active_paints_to_machine.insert_one(active_paint.model_dump())
    
    # Stoktan düş
    new_stock = current_stock - given_amount_kg
    await db.paints.update_one({"id": paint_id}, {"$set": {"stock_kg": new_stock}})
    
    # Hareket kaydı
    movement = PaintMovement(
        paint_id=paint_id,
        paint_name=paint["name"],
        movement_type="to_machine",
        amount_kg=given_amount_kg,
        machine_id=machine_id,
        machine_name=machine_name,
        note=f"Makineye verildi: {given_amount_kg} kg"
    )
    await db.paint_movements.insert_one(movement.model_dump())
    
    return {
        "message": f"{paint['name']} - {given_amount_kg} kg {machine_name} makinesine verildi",
        "active_paint_id": active_paint.id,
        "new_stock": new_stock
    }

@api_router.get("/paints/active-on-machines")
async def get_active_paints_on_machines():
    """Makinelerde aktif (geri alınmamış) boyaları listele"""
    active_paints = await db.active_paints_to_machine.find(
        {"returned": False},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return active_paints

@api_router.post("/paints/return-from-machine")
async def return_paint_from_machine(data: dict = Body(...)):
    """Makineden boya geri al - tartıdan okunan değeri gir, sistem farkı hesaplar"""
    active_paint_id = data.get("active_paint_id")
    returned_amount_kg = float(data.get("returned_amount_kg", 0))
    
    if returned_amount_kg < 0:
        raise HTTPException(status_code=400, detail="Geçersiz miktar")
    
    # Aktif boya kaydını bul
    active_paint = await db.active_paints_to_machine.find_one(
        {"id": active_paint_id, "returned": False},
        {"_id": 0}
    )
    if not active_paint:
        raise HTTPException(status_code=404, detail="Aktif boya kaydı bulunamadı")
    
    given_amount = active_paint["given_amount_kg"]
    used_amount = given_amount - returned_amount_kg
    
    if used_amount < 0:
        raise HTTPException(status_code=400, detail="Geri alınan miktar verilen miktardan fazla olamaz")
    
    # Aktif kaydı güncelle
    await db.active_paints_to_machine.update_one(
        {"id": active_paint_id},
        {"$set": {
            "returned": True,
            "returned_amount_kg": returned_amount_kg,
            "used_amount_kg": used_amount,
            "returned_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Geri alınan miktarı stoka ekle
    paint = await db.paints.find_one({"id": active_paint["paint_id"]}, {"_id": 0})
    if paint:
        new_stock = paint.get("stock_kg", 0) + returned_amount_kg
        await db.paints.update_one({"id": active_paint["paint_id"]}, {"$set": {"stock_kg": new_stock}})
    
    # Hareket kaydı - Geri alındı
    movement_return = PaintMovement(
        paint_id=active_paint["paint_id"],
        paint_name=active_paint["paint_name"],
        movement_type="from_machine",
        amount_kg=returned_amount_kg,
        machine_id=active_paint["machine_id"],
        machine_name=active_paint["machine_name"],
        note=f"Makineden geri alındı: {returned_amount_kg} kg"
    )
    await db.paint_movements.insert_one(movement_return.model_dump())
    
    # Hareket kaydı - Kullanılan
    if used_amount > 0:
        movement_used = PaintMovement(
            paint_id=active_paint["paint_id"],
            paint_name=active_paint["paint_name"],
            movement_type="used",
            amount_kg=used_amount,
            machine_id=active_paint["machine_id"],
            machine_name=active_paint["machine_name"],
            note=f"Makine kullanımı: {used_amount} kg (Verilen: {given_amount} kg, Kalan: {returned_amount_kg} kg)"
        )
        await db.paint_movements.insert_one(movement_used.model_dump())
    
    return {
        "message": f"{active_paint['paint_name']} - {active_paint['machine_name']} makinesinden geri alındı",
        "given_amount": given_amount,
        "returned_amount": returned_amount_kg,
        "used_amount": used_amount,
        "new_stock": paint.get("stock_kg", 0) + returned_amount_kg if paint else 0
    }

# ==================== KULLANICI YÖNETİMİ ====================

@api_router.post("/users", response_model=User)
async def create_user(data: dict = Body(...)):
    """Yeni kullanıcı oluştur"""
    username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "")
    display_name = data.get("display_name", username)
    phone = data.get("phone", "")
    
    if not username or not password or not role:
        raise HTTPException(status_code=400, detail="Kullanıcı adı, şifre ve rol zorunludur")
    
    if role not in ["operator", "plan", "depo", "sofor"]:
        raise HTTPException(status_code=400, detail="Geçersiz rol")
    
    # Kullanıcı adı kontrolü
    existing = await db.users.find_one({"username": username, "is_active": True})
    if existing:
        raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten kullanılıyor")
    
    user = User(
        username=username,
        password=password,
        role=role,
        display_name=display_name,
        phone=phone
    )
    await db.users.insert_one(user.model_dump())
    await log_audit("Yonetim", "create", "user", username, f"Rol: {role}")

    
    # Şifreyi döndürmeden önce kaldır
    user_dict = user.model_dump()
    user_dict.pop("password", None)
    return user

@api_router.get("/users")
async def get_users(role: Optional[str] = None):
    """Kullanıcıları listele"""
    query = {"is_active": True}
    if role:
        query["role"] = role
    users = await db.users.find(query, {"_id": 0, "password": 0}).sort("created_at", -1).to_list(200)
    return users

@api_router.post("/users/login")
async def user_login(data: dict = Body(...)):
    """Kullanıcı girişi (rol bazlı)"""
    username = data.get("username", "").strip()
    password = data.get("password", "")
    expected_role = data.get("role")  # Hangi sayfadan giriş yapılıyor
    
    user = await db.users.find_one({
        "username": username,
        "password": password,
        "is_active": True
    }, {"_id": 0})
    
    if not user:
        raise HTTPException(status_code=401, detail="Geçersiz kullanıcı adı veya şifre")
    
    # Rol kontrolü
    if expected_role and user["role"] != expected_role:
        raise HTTPException(status_code=403, detail=f"Bu sayfaya erişim yetkiniz yok. Yetkiniz: {user['role']}")
    
    # Şifreyi döndürme
    user.pop("password", None)
    return user

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str):
    """Kullanıcı sil (pasif yap)"""
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"is_active": False}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return {"success": True}

@api_router.put("/users/{user_id}/location")
async def update_user_location(user_id: str, data: dict = Body(...)):
    """Kullanıcı konumunu güncelle (şoförler için)"""
    lat = data.get("lat")
    lng = data.get("lng")
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "current_location_lat": lat,
            "current_location_lng": lng,
            "location_updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"success": True}

@api_router.get("/users/drivers/locations")
async def get_driver_locations():
    """Tüm şoförlerin konumlarını getir"""
    drivers = await db.users.find({
        "role": "sofor",
        "is_active": True,
        "current_location_lat": {"$ne": None}
    }, {"_id": 0, "password": 0}).to_list(100)
    return drivers

@api_router.get("/paints/analytics")
async def get_paint_analytics(period: str = "weekly"):
    """Boya tüketim analitiği - Gerçek kullanım (makineden geri alındıktan sonra hesaplanan fark)"""
    from datetime import timedelta
    
    if period == "weekly":
        date_ago = datetime.now(timezone.utc) - timedelta(days=7)
    else:
        date_ago = datetime.now(timezone.utc) - timedelta(days=30)
    
    date_ago_str = date_ago.isoformat()
    
    # Gerçek kullanım hareketleri ("used" = makineden geri alındığında hesaplanan fark)
    # ve "remove" (stoktan direkt çıkarılan)
    movements = await db.paint_movements.find(
        {
            "movement_type": {"$in": ["remove", "used"]},
            "created_at": {"$gte": date_ago_str}
        },
        {"_id": 0}
    ).to_list(1000)
    
    # Boya bazında tüketim
    paint_consumption = {}
    machine_consumption = {}
    daily_consumption = {}
    
    for mov in movements:
        paint_name = mov["paint_name"]
        amount = mov["amount_kg"]
        machine_name = mov.get("machine_name", "Bilinmeyen")
        
        # Boya bazında
        if paint_name not in paint_consumption:
            paint_consumption[paint_name] = 0
        paint_consumption[paint_name] += amount
        
        # Makine bazında (sadece "used" tipi için)
        if mov["movement_type"] == "used" and machine_name and machine_name != "Bilinmeyen":
            if machine_name not in machine_consumption:
                machine_consumption[machine_name] = 0
            machine_consumption[machine_name] += amount
        
        # Günlük bazda
        date_str = mov["created_at"][:10]
        if date_str not in daily_consumption:
            daily_consumption[date_str] = 0
        daily_consumption[date_str] += amount
    
    return {
        "period": period,
        "paint_consumption": paint_consumption,
        "machine_consumption": machine_consumption,
        "daily_consumption": daily_consumption,
        "total_consumed": sum(paint_consumption.values())
    }

@api_router.get("/paints/low-stock")
async def get_low_stock_paints():
    """Düşük stoklu boyaları listele (5L altı)"""
    paints = await db.paints.find({"stock_kg": {"$lt": LOW_STOCK_THRESHOLD}}, {"_id": 0}).to_list(100)
    return {
        "threshold": LOW_STOCK_THRESHOLD,
        "low_stock_paints": paints
    }

# ==================== MAKİNE MESAJ SİSTEMİ ====================

@api_router.post("/messages", response_model=MachineMessage)
async def send_message(data: dict = Body(...)):
    """Plan veya Yönetim'den makineye mesaj gönder"""
    machine_id = data.get("machine_id")
    machine_name = data.get("machine_name")
    sender_role = data.get("sender_role")  # "yonetim" veya "plan"
    sender_name = data.get("sender_name", sender_role.title())
    message_text = data.get("message")
    
    if not all([machine_id, sender_role, message_text]):
        raise HTTPException(status_code=400, detail="Eksik bilgi")
    
    message = MachineMessage(
        machine_id=machine_id,
        machine_name=machine_name or "",
        sender_role=sender_role,
        sender_name=sender_name,
        message=message_text
    )
    
    await db.machine_messages.insert_one(message.model_dump())
    
    # WebSocket ile operatöre bildirim gönder
    await manager.broadcast({
        "type": "new_message",
        "data": {
            "machine_id": machine_id,
            "machine_name": machine_name,
            "sender_role": sender_role,
            "sender_name": sender_name,
            "message": message_text,
            "created_at": message.created_at
        }
    })
    
    # FCM Push Bildirimi gönder (operatörlere)
    try:
        await send_notification_to_operators(
            machine_id=machine_id,
            title=f"📩 Yeni Mesaj - {sender_name}",
            body=message_text[:100],
            data={"type": "new_message", "machine_id": machine_id}
        )
    except Exception as e:
        logging.error(f"FCM notification error for message: {e}")
    
    return message

@api_router.get("/messages/{machine_id}")
async def get_machine_messages(machine_id: str, limit: int = 50):
    """Bir makinenin mesajlarını getir"""
    messages = await db.machine_messages.find(
        {"machine_id": machine_id}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    return list(reversed(messages))  # En eski en üstte

@api_router.get("/messages/{machine_id}/unread")
async def get_unread_messages(machine_id: str):
    """Okunmamış mesaj sayısını getir"""
    count = await db.machine_messages.count_documents({
        "machine_id": machine_id,
        "is_read": False
    })
    return {"unread_count": count}

@api_router.put("/messages/{machine_id}/mark-read")
async def mark_messages_read(machine_id: str):
    """Bir makinenin tüm mesajlarını okundu olarak işaretle"""
    result = await db.machine_messages.update_many(
        {"machine_id": machine_id, "is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"marked_read": result.modified_count}

@api_router.get("/messages/all/incoming")
async def get_all_incoming_messages(limit: int = 100):
    """Tüm operatör mesajlarını getir (Yönetim için)"""
    messages = await db.machine_messages.find(
        {"sender_role": "operator"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(limit)
    return messages

@api_router.get("/messages/all/unread-count")
async def get_all_unread_count():
    """Tüm okunmamış operatör mesaj sayısı"""
    count = await db.machine_messages.count_documents({
        "sender_role": "operator",
        "is_read": False
    })
    return {"unread_count": count}

@api_router.put("/messages/mark-read/{message_id}")
async def mark_single_message_read(message_id: str):
    """Tek bir mesajı okundu olarak işaretle"""
    result = await db.machine_messages.update_one(
        {"id": message_id},
        {"$set": {"is_read": True}}
    )
    return {"success": result.modified_count > 0}

# ==================== ZİYARETÇİ TAKİP ====================

def parse_user_agent(user_agent: str) -> dict:
    """User agent string'inden cihaz bilgisi çıkar"""
    ua = user_agent.lower()
    
    # Cihaz tipi
    if "mobile" in ua or "android" in ua and "mobile" in ua:
        device_type = "Mobil"
    elif "tablet" in ua or "ipad" in ua:
        device_type = "Tablet"
    else:
        device_type = "Masaüstü"
    
    # İşletim sistemi
    if "windows" in ua:
        os = "Windows"
    elif "mac" in ua or "macintosh" in ua:
        os = "MacOS"
    elif "iphone" in ua:
        os = "iOS"
        device_type = "Mobil"
    elif "ipad" in ua:
        os = "iOS"
        device_type = "Tablet"
    elif "android" in ua:
        os = "Android"
    elif "linux" in ua:
        os = "Linux"
    else:
        os = "Bilinmeyen"
    
    # Tarayıcı
    if "chrome" in ua and "edg" not in ua:
        browser = "Chrome"
    elif "firefox" in ua:
        browser = "Firefox"
    elif "safari" in ua and "chrome" not in ua:
        browser = "Safari"
    elif "edg" in ua:
        browser = "Edge"
    elif "opera" in ua or "opr" in ua:
        browser = "Opera"
    else:
        browser = "Diğer"
    
    # Cihaz modeli tahmini
    if "iphone" in ua:
        device_model = "iPhone"
    elif "ipad" in ua:
        device_model = "iPad"
    elif "samsung" in ua:
        device_model = "Samsung"
    elif "huawei" in ua:
        device_model = "Huawei"
    elif "xiaomi" in ua or "redmi" in ua:
        device_model = "Xiaomi"
    elif "pixel" in ua:
        device_model = "Google Pixel"
    elif "windows" in ua:
        device_model = "PC"
    elif "macintosh" in ua:
        device_model = "Mac"
    else:
        device_model = device_type
    
    return {
        "device_type": device_type,
        "device_model": device_model,
        "browser": browser,
        "os": os
    }

from fastapi import Request

@api_router.post("/visitors/log")
async def log_visitor(request: Request, data: dict = Body(...)):
    """Ziyaretçi kaydı oluştur"""
    # IP adresini al
    forwarded_for = request.headers.get("x-forwarded-for")
    if forwarded_for:
        ip_address = forwarded_for.split(",")[0].strip()
    else:
        ip_address = request.client.host if request.client else "Bilinmeyen"
    
    user_agent = data.get("user_agent", "")
    page_visited = data.get("page_visited", "/")
    
    # User agent'ı parse et
    device_info = parse_user_agent(user_agent)
    
    visitor = Visitor(
        ip_address=ip_address,
        user_agent=user_agent,
        device_type=device_info["device_type"],
        device_model=device_info["device_model"],
        browser=device_info["browser"],
        os=device_info["os"],
        page_visited=page_visited
    )
    
    await db.visitors.insert_one(visitor.model_dump())
    return {"message": "Ziyaret kaydedildi", "visitor_id": visitor.id}

@api_router.get("/visitors")
async def get_visitors(limit: int = 100):
    """Ziyaretçi listesini getir"""
    visitors = await db.visitors.find({}, {"_id": 0}).sort("visited_at", -1).to_list(limit)
    return visitors

@api_router.get("/visitors/stats")
async def get_visitor_stats():
    """Ziyaretçi istatistikleri"""
    total = await db.visitors.count_documents({})
    
    # Son 24 saat
    from datetime import timedelta
    day_ago = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
    today_count = await db.visitors.count_documents({"visited_at": {"$gte": day_ago}})
    
    # Son 7 gün
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    week_count = await db.visitors.count_documents({"visited_at": {"$gte": week_ago}})
    
    # Cihaz dağılımı
    device_pipeline = [
        {"$group": {"_id": "$device_type", "count": {"$sum": 1}}}
    ]
    device_stats = await db.visitors.aggregate(device_pipeline).to_list(10)
    
    return {
        "total_visitors": total,
        "today": today_count,
        "this_week": week_count,
        "device_distribution": {item["_id"]: item["count"] for item in device_stats}
    }

@api_router.delete("/visitors/clear")
async def clear_visitors():
    """Tüm ziyaretçi kayıtlarını temizle"""
    result = await db.visitors.delete_many({})
    return {"message": f"{result.deleted_count} kayıt silindi"}

# ==================== OPERATÖR OTURUM YÖNETİMİ ====================

@api_router.post("/operator/session")
async def create_or_get_session(data: dict = Body(...)):
    """Operatör oturumu oluştur veya mevcut oturumu getir"""
    device_id = data.get("device_id")
    operator_name = data.get("operator_name")
    machine_id = data.get("machine_id")
    machine_name = data.get("machine_name")
    
    if not device_id:
        raise HTTPException(status_code=400, detail="device_id gerekli")
    
    # Bugünün sonunu hesapla
    today_end = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59)
    
    # Mevcut oturumu kontrol et
    existing = await db.operator_sessions.find_one({
        "device_id": device_id,
        "expires_at": {"$gt": datetime.now(timezone.utc).isoformat()}
    }, {"_id": 0})
    
    if existing and not operator_name:
        # Mevcut oturum var, güncelle ve döndür
        await db.operator_sessions.update_one(
            {"id": existing["id"]},
            {"$set": {"last_active": datetime.now(timezone.utc).isoformat()}}
        )
        return existing
    
    if operator_name:
        # Yeni oturum oluştur veya güncelle
        session = OperatorSession(
            device_id=device_id,
            operator_name=operator_name,
            machine_id=machine_id,
            machine_name=machine_name,
            expires_at=today_end.isoformat()
        )
        
        # Eski oturumu sil ve yenisini ekle
        await db.operator_sessions.delete_many({"device_id": device_id})
        await db.operator_sessions.insert_one(session.model_dump())
        return session.model_dump()
    
    return None

@api_router.get("/operator/session/{device_id}")
async def get_operator_session(device_id: str):
    """Cihaz ID'sine göre operatör oturumunu getir"""
    session = await db.operator_sessions.find_one({
        "device_id": device_id,
        "expires_at": {"$gt": datetime.now(timezone.utc).isoformat()}
    }, {"_id": 0})
    return session

@api_router.delete("/operator/session/{device_id}")
async def delete_operator_session(device_id: str):
    """Operatör oturumunu sonlandır"""
    result = await db.operator_sessions.delete_many({"device_id": device_id})
    return {"deleted": result.deleted_count}

# ==================== PALET YÖNETİMİ ====================

@api_router.post("/pallets")
async def create_pallet(data: dict = Body(...)):
    """Yeni palet oluştur/tara - hem PalletScan hem Pallet formatını destekler"""
    # PalletScan formatı (WarehouseFlow - pallet_code kullanır)
    pallet_code = data.get("pallet_code") or data.get("code", "")
    
    if data.get("machine_id"):
        # Tam Pallet formatı (machine bilgisi var)
        pallet = Pallet(
            code=pallet_code,
            job_id=data.get("job_id", ""),
            job_name=data.get("job_name", ""),
            machine_id=data.get("machine_id", ""),
            machine_name=data.get("machine_name", ""),
            koli_count=data.get("koli_count", 0),
            operator_name=data.get("operator_name", "")
        )
        doc = pallet.model_dump()
        await db.pallets.insert_one(doc)
        return {k: v for k, v in doc.items() if k != "_id"}
        await log_audit(data.get("operator_name", "Depo"), "create", "pallet", pallet_code, f"Is: {data.get('job_name', '')}")

    else:
        # PalletScan formatı (basit tarama)
        scan = PalletScan(
            pallet_code=pallet_code,
            job_id=data.get("job_id", "unknown"),
            job_name=data.get("job_name", ""),
            operator_name=data.get("operator_name", "")
        )
        doc = scan.model_dump()
        await db.pallets.insert_one(doc)
        return {k: v for k, v in doc.items() if k != "_id"}
        await log_audit(data.get("operator_name", "Depo"), "create", "pallet", pallet_code, f"Tarama - Is: {data.get('job_name', '')}")


@api_router.get("/pallets")
async def get_pallets(job_id: Optional[str] = None, status: Optional[str] = None):
    """Paletleri listele"""
    query = {}
    if job_id:
        query["job_id"] = job_id
    if status:
        query["status"] = status
    pallets = await db.pallets.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return pallets

@api_router.get("/pallets/by-job/{job_id}")
async def get_pallets_by_job(job_id: str):
    """Bir işe ait paletleri getir"""
    pallets = await db.pallets.find({"job_id": job_id}, {"_id": 0}).to_list(100)
    total_koli = sum(p["koli_count"] for p in pallets)
    return {
        "pallets": pallets,
        "total_pallets": len(pallets),
        "total_koli": total_koli
    }

@api_router.get("/pallets/search")
async def search_pallets(q: str):
    """Palet ara (kod veya iş adı ile)"""
    pallets = await db.pallets.find({
        "$or": [
            {"code": {"$regex": q, "$options": "i"}},
            {"job_name": {"$regex": q, "$options": "i"}}
        ]
    }, {"_id": 0}).to_list(50)
    return pallets

@api_router.put("/pallets/{pallet_id}/status")
async def update_pallet_status(pallet_id: str, data: dict = Body(...)):
    """Palet durumunu güncelle"""
    status = data.get("status")
    await db.pallets.update_one(
        {"id": pallet_id},
        {"$set": {"status": status}}
    )
    return {"success": True}

# ==================== ARAÇ YÖNETİMİ ====================

@api_router.post("/vehicles", response_model=Vehicle)
async def create_vehicle(data: dict = Body(...)):
    """Yeni araç ekle"""
    vehicle = Vehicle(
        plate=data.get("plate"),
        driver_name=data.get("driver_name")
    )
    await db.vehicles.insert_one(vehicle.model_dump())
    return vehicle

@api_router.get("/vehicles")
async def get_vehicles():
    """Araçları listele"""
    vehicles = await db.vehicles.find({"is_active": True}, {"_id": 0}).to_list(100)
    return vehicles

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str):
    """Araç sil (pasif yap)"""
    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"is_active": False}}
    )
    return {"success": True}

# ==================== SEVKİYAT YÖNETİMİ ====================

@api_router.post("/shipments", response_model=Shipment)
async def create_shipment(data: dict = Body(...)):
    """Yeni sevkiyat oluştur"""
    pallet_ids = data.get("pallet_ids", [])
    
    # Paletlerin toplam koli sayısını hesapla
    total_koli = 0
    if pallet_ids:
        pallets = await db.pallets.find({"id": {"$in": pallet_ids}}, {"_id": 0}).to_list(100)
        total_koli = sum(p["koli_count"] for p in pallets)
        # Paletleri sevkiyata ata
        await db.pallets.update_many(
            {"id": {"$in": pallet_ids}},
            {"$set": {"status": "in_shipment"}}
        )
    
    shipment = Shipment(
        vehicle_id=data.get("vehicle_id"),
        vehicle_plate=data.get("vehicle_plate"),
        driver_id=data.get("driver_id"),
        driver_name=data.get("driver_name"),
        pallets=pallet_ids,
        total_koli=data.get("total_koli", total_koli),
        delivery_address=data.get("delivery_address"),
        delivery_phone=data.get("delivery_phone"),
        delivery_notes=data.get("delivery_notes"),
        created_by=data.get("created_by", "plan")
    )
    await db.shipments.insert_one(shipment.model_dump())
    return shipment

@api_router.get("/shipments")
async def get_shipments(status: Optional[str] = None, driver_id: Optional[str] = None):
    """Sevkiyatları listele"""
    query = {}
    if status:
        query["status"] = status
    if driver_id:
        query["driver_id"] = driver_id
    shipments = await db.shipments.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return shipments

@api_router.get("/shipments/{shipment_id}")
async def get_shipment(shipment_id: str):
    """Sevkiyat detayını getir"""
    shipment = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
    if not shipment:
        raise HTTPException(status_code=404, detail="Sevkiyat bulunamadı")
    
    # Paletleri de getir
    pallets = await db.pallets.find({"id": {"$in": shipment.get("pallets", [])}}, {"_id": 0}).to_list(100)
    shipment["pallet_details"] = pallets
    return shipment

@api_router.put("/shipments/{shipment_id}")
async def update_shipment(shipment_id: str, data: dict = Body(...)):
    """Sevkiyat güncelle"""
    update_data = {}
    for key in ["delivery_address", "delivery_phone", "delivery_notes", "total_koli", "driver_id", "driver_name", "vehicle_id", "vehicle_plate"]:
        if key in data:
            update_data[key] = data[key]
    
    if "pallet_ids" in data:
        update_data["pallets"] = data["pallet_ids"]
        # Paletleri güncelle
        await db.pallets.update_many(
            {"id": {"$in": data["pallet_ids"]}},
            {"$set": {"status": "in_shipment"}}
        )
    
    await db.shipments.update_one(
        {"id": shipment_id},
        {"$set": update_data}
    )
    return {"success": True}

@api_router.put("/shipments/{shipment_id}/status")
async def update_shipment_status(shipment_id: str, data: dict = Body(...)):
    """Sevkiyat durumunu güncelle (Şoför için)"""
    status = data.get("status")
    reason = data.get("reason")
    
    update_data = {"status": status}
    if status == "in_transit":
        update_data["started_at"] = datetime.now(timezone.utc).isoformat()
    elif status in ["delivered", "failed"]:
        update_data["completed_at"] = datetime.now(timezone.utc).isoformat()
        if reason:
            update_data["delivery_status_reason"] = reason
        
        # Paletleri güncelle
        shipment = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
        if shipment:
            pallet_status = "delivered" if status == "delivered" else "in_warehouse"
            await db.pallets.update_many(
                {"id": {"$in": shipment.get("pallets", [])}},
                {"$set": {"status": pallet_status}}
            )
    
    await db.shipments.update_one(
        {"id": shipment_id},
        {"$set": update_data}
    )
    return {"success": True}

@api_router.delete("/shipments/{shipment_id}")
async def delete_shipment(shipment_id: str):
    """Sevkiyat sil"""
    shipment = await db.shipments.find_one({"id": shipment_id}, {"_id": 0})
    if shipment:
        # Paletleri depoya geri al
        await db.pallets.update_many(
            {"id": {"$in": shipment.get("pallets", [])}},
            {"$set": {"status": "in_warehouse"}}
        )
    await db.shipments.delete_one({"id": shipment_id})
    return {"success": True}

# ==================== ŞOFÖR YÖNETİMİ ====================

@api_router.post("/drivers", response_model=Driver)
async def create_driver(data: dict = Body(...)):
    """Yeni şoför ekle"""
    driver = Driver(
        name=data.get("name"),
        password=data.get("password"),
        phone=data.get("phone")
    )
    await db.drivers.insert_one(driver.model_dump())
    return driver

@api_router.get("/drivers")
async def get_drivers():
    """Şoförleri listele"""
    drivers = await db.drivers.find({"is_active": True}, {"_id": 0, "password": 0}).to_list(100)
    return drivers

@api_router.post("/drivers/login")
async def driver_login(data: dict = Body(...)):
    """Şoför girişi"""
    name = data.get("name")
    password = data.get("password")
    
    driver = await db.drivers.find_one({
        "name": name,
        "password": password,
        "is_active": True
    }, {"_id": 0})
    
    if not driver:
        raise HTTPException(status_code=401, detail="Geçersiz kullanıcı adı veya şifre")
    
    # Şifreyi döndürme
    driver.pop("password", None)
    return driver

@api_router.put("/drivers/{driver_id}/location")
async def update_driver_location(driver_id: str, data: dict = Body(...)):
    """Şoför konumunu güncelle"""
    lat = data.get("lat")
    lng = data.get("lng")
    
    await db.drivers.update_one(
        {"id": driver_id},
        {"$set": {
            "current_location_lat": lat,
            "current_location_lng": lng,
            "location_updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"success": True}

@api_router.get("/drivers/{driver_id}/location")
async def get_driver_location(driver_id: str):
    """Şoför konumunu getir"""
    driver = await db.drivers.find_one({"id": driver_id}, {"_id": 0, "password": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Şoför bulunamadı")
    return {
        "lat": driver.get("current_location_lat"),
        "lng": driver.get("current_location_lng"),
        "updated_at": driver.get("location_updated_at")
    }

@api_router.get("/drivers/{driver_id}/shipments")
async def get_driver_shipments(driver_id: str):
    """Şoföre atanan sevkiyatları getir"""
    shipments = await db.shipments.find({
        "driver_id": driver_id,
        "status": {"$in": ["preparing", "in_transit"]}
    }, {"_id": 0}).sort("created_at", -1).to_list(50)
    return shipments

# ==================== DEPO SEVKİYAT KAYDI ====================

@api_router.post("/warehouse/shipment-log")
async def create_warehouse_shipment_log(data: dict = Body(...)):
    """Depo sevkiyat kaydı oluştur"""
    log = WarehouseShipmentLog(
        shipment_id=data.get("shipment_id"),
        vehicle_plate=data.get("vehicle_plate"),
        pallet_ids=data.get("pallet_ids", []),
        total_koli=data.get("total_koli", 0),
        partial=data.get("partial", False),
        delivered_koli=data.get("delivered_koli", 0),
        operator_name=data.get("operator_name", "Depo")
    )
    await db.warehouse_shipment_logs.insert_one(log.model_dump())
    return log

@api_router.get("/warehouse/shipment-logs")
async def get_warehouse_shipment_logs(limit: int = 100):
    """Depo sevkiyat kayıtlarını listele"""
    logs = await db.warehouse_shipment_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return logs

@api_router.get("/analytics/daily-by-week")
async def get_daily_analytics_by_week(week_offset: int = 0):
    """Hafta bazında günlük üretim analitiği - Vardiya sonu raporları dahil"""
    from datetime import timedelta
    
    # Hedef haftayı hesapla
    today = datetime.now(timezone.utc)
    # Pazartesi'yi bul
    days_since_monday = today.weekday()
    this_monday = today - timedelta(days=days_since_monday)
    target_monday = this_monday + timedelta(weeks=week_offset)
    target_sunday = target_monday + timedelta(days=6)
    
    daily_stats = []
    for i in range(7):
        date = target_monday + timedelta(days=i)
        start_of_day = date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        
        # Tamamlanan işler
        jobs = await db.jobs.find(
            {"status": "completed", "completed_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)
        
        # Vardiya sonu raporları (kısmi üretimler)
        shift_reports = await db.shift_end_reports.find(
            {"created_at": {"$gte": start_of_day.isoformat(), "$lt": end_of_day.isoformat()}},
            {"_id": 0}
        ).to_list(1000)
        
        total_koli = sum(job.get("completed_koli", job.get("koli_count", 0)) for job in jobs)
        machine_breakdown = {}
        
        for job in jobs:
            machine = job["machine_name"]
            if machine not in machine_breakdown:
                machine_breakdown[machine] = 0
            machine_breakdown[machine] += job.get("completed_koli", job.get("koli_count", 0))
        
        # Vardiya raporlarından kısmi üretim ekle (tamamlanmamış işler için) - Batch query
        job_ids = [r["job_id"] for r in shift_reports if r.get("job_id") and r.get("produced_koli", 0) > 0]
        if job_ids:
            jobs_list = await db.jobs.find({"id": {"$in": job_ids}}, {"_id": 0}).to_list(1000)
            jobs_dict = {j["id"]: j for j in jobs_list}
            
            for report in shift_reports:
                produced = report.get("produced_koli", 0)
                if produced > 0:
                    job_id = report.get("job_id")
                    if job_id:
                        job = jobs_dict.get(job_id)
                        if job and job.get("status") == "completed":
                            continue
                    
                    machine = report.get("machine_name", "")
                    if machine:
                        if machine not in machine_breakdown:
                            machine_breakdown[machine] = 0
                        machine_breakdown[machine] += produced
                        total_koli += produced
        
        day_names = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
        daily_stats.append({
            "date": start_of_day.strftime("%d %b"),
            "day_name": day_names[i],
            "full_date": start_of_day.strftime("%Y-%m-%d"),
            "total_koli": total_koli,
            "machines": machine_breakdown
        })
    
    return {
        "week_start": target_monday.strftime("%d %b %Y"),
        "week_end": target_sunday.strftime("%d %b %Y"),
        "week_offset": week_offset,
        "daily_stats": daily_stats
    }

# Yönetici kaydı endpoint
@api_router.post("/managers/register")
async def register_manager(data: dict = Body(...)):
    """Yöneticiyi bildirim için kaydet"""
    manager_id = data.get("manager_id")
    if not manager_id:
        raise HTTPException(status_code=400, detail="manager_id required")
    
    # Yönetici kaydını veritabanına ekle (opsiyonel, aktif bağlantılar ws üzerinden takip ediliyor)
    await db.active_managers.update_one(
        {"manager_id": manager_id},
        {"$set": {"manager_id": manager_id, "registered_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    logging.info(f"Manager registered: {manager_id}")
    return {"status": "registered", "manager_id": manager_id}

# FCM Token kaydetme endpoint
@api_router.post("/notifications/register-token")
async def register_fcm_token(data: dict = Body(...)):
    """FCM token'ı kaydet"""
    token = data.get("token")
    user_type = data.get("user_type", "manager")  # manager, operator, plan
    user_id = data.get("user_id", "")
    
    if not token:
        raise HTTPException(status_code=400, detail="token required")
    
    await db.fcm_tokens.update_one(
        {"token": token},
        {"$set": {
            "token": token,
            "user_type": user_type,
            "user_id": user_id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    logging.info(f"FCM token registered for {user_type}: {token[:20]}...")
    return {"status": "registered"}

# Tüm yöneticilere bildirim gönderme (internal function)
async def send_notification_to_managers(title: str, body: str, data: dict = None):
    """Tüm kayıtlı yöneticilere FCM bildirimi gönder"""
    try:
        # Yönetici token'larını al
        tokens_cursor = db.fcm_tokens.find({"user_type": "manager"}, {"token": 1, "_id": 0})
        tokens = [doc["token"] async for doc in tokens_cursor]
        
        if tokens:
            await send_fcm_notification(tokens, title, body, data)
            logging.info(f"Notification sent to {len(tokens)} managers")
        else:
            logging.warning("No manager FCM tokens found")
    except Exception as e:
        logging.error(f"Error sending notification to managers: {e}")

# Belirli makinedeki operatörlere bildirim gönderme
async def send_notification_to_operators(machine_id: str, title: str, body: str, data: dict = None):
    """Belirli bir makinedeki operatörlere FCM bildirimi gönder"""
    try:
        # Tüm operatör token'larını al (şimdilik hepsine gönder)
        tokens_cursor = db.fcm_tokens.find({"user_type": "operator"}, {"token": 1, "_id": 0})
        tokens = [doc["token"] async for doc in tokens_cursor]
        
        if tokens:
            await send_fcm_notification(tokens, title, body, data)
            logging.info(f"Notification sent to {len(tokens)} operators for machine {machine_id}")
        else:
            logging.warning("No operator FCM tokens found")
    except Exception as e:
        logging.error(f"Error sending notification to operators: {e}")

# Plan kullanıcılarına bildirim gönderme
async def send_notification_to_plan_users(title: str, body: str, data: dict = None):
    """Tüm kayıtlı Plan kullanıcılarına FCM bildirimi gönder"""
    try:
        tokens_cursor = db.fcm_tokens.find({"user_type": "plan"}, {"token": 1, "_id": 0})
        tokens = [doc["token"] async for doc in tokens_cursor]
        
        if tokens:
            await send_fcm_notification(tokens, title, body, data)
            logging.info(f"Notification sent to {len(tokens)} plan users")
        else:
            logging.warning("No plan FCM tokens found")
    except Exception as e:
        logging.error(f"Error sending notification to plan users: {e}")

# Tüm çalışanlara bildirim gönderme (vardiya başlat/bitir için)
async def send_notification_to_all_workers(title: str, body: str, data: dict = None):
    """Tüm operatör ve plan kullanıcılarına FCM bildirimi gönder"""
    try:
        tokens_cursor = db.fcm_tokens.find(
            {"user_type": {"$in": ["operator", "plan"]}}, 
            {"token": 1, "_id": 0}
        )
        tokens = [doc["token"] async for doc in tokens_cursor]
        
        if tokens:
            await send_fcm_notification(tokens, title, body, data)
            logging.info(f"Notification sent to {len(tokens)} workers")
        else:
            logging.warning("No worker FCM tokens found")
    except Exception as e:
        logging.error(f"Error sending notification to all workers: {e}")

app.include_router(api_router)

# WebSocket endpoint - Yönetici bildirimleri için
@app.websocket("/api/ws/manager/{manager_id}")
async def manager_websocket(websocket: WebSocket, manager_id: str):
    await manager_ws.connect(websocket, manager_id)
    logging.info(f"Manager WebSocket connected: {manager_id}")
    try:
        while True:
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager_ws.disconnect(manager_id)
        logging.info(f"Manager WebSocket disconnected: {manager_id}")
    except Exception as e:
        logging.error(f"Manager WebSocket error: {e}")
        manager_ws.disconnect(manager_id)

# WebSocket endpoint - Depo bildirimleri için
@app.websocket("/api/ws/warehouse")
async def warehouse_websocket(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Bağlantıyı canlı tutmak için ping-pong
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception as e:
        logging.error(f"WebSocket error: {e}")
        manager.disconnect(websocket)

# WebSocket endpoint - Operatör bildirimleri için
@app.websocket("/api/ws/operator/{machine_id}")
async def operator_websocket(websocket: WebSocket, machine_id: str):
    await manager.connect(websocket)
    logging.info(f"Operator WebSocket connected for machine: {machine_id}")
    try:
        while True:
            # Bağlantıyı canlı tutmak için ping-pong
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect(websocket)
        logging.info(f"Operator WebSocket disconnected for machine: {machine_id}")
    except Exception as e:
        logging.error(f"Operator WebSocket error: {e}")
        manager.disconnect(websocket)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
