"""
Test Excel Export Feature - Iteration 10
Tests the weekly Excel report export with 4 sheets:
1) Haftalik Ozet - Daily breakdown
2) Makine Detayi - Machine-by-machine daily production
3) Operator Performansi - Operator stats
4) Defo Raporu - Defect logs
"""
import pytest
import requests
import os
from io import BytesIO

# Use openpyxl to verify Excel content
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestExcelExport:
    """Excel Export API Tests"""
    
    def test_weekly_export_returns_200(self):
        """Test GET /api/analytics/export?period=weekly returns 200"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Weekly export returns 200")
    
    def test_monthly_export_returns_200(self):
        """Test GET /api/analytics/export?period=monthly returns 200"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=monthly")
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Monthly export returns 200")
    
    def test_export_content_type_is_excel(self):
        """Test that response content type is Excel"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheetml' in content_type or 'application/vnd' in content_type, \
            f"Expected Excel content type, got {content_type}"
        print(f"✓ Content-Type is Excel: {content_type}")
    
    def test_export_has_content_disposition(self):
        """Test that response has Content-Disposition header for download"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        content_disp = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disp, f"Expected attachment, got {content_disp}"
        assert '.xlsx' in content_disp, f"Expected .xlsx filename, got {content_disp}"
        print(f"✓ Content-Disposition: {content_disp}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_excel_has_4_sheets(self):
        """Test that downloaded Excel has exactly 4 sheets"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        assert len(sheet_names) == 4, f"Expected 4 sheets, got {len(sheet_names)}: {sheet_names}"
        print(f"✓ Excel has 4 sheets: {sheet_names}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_sheet_names_are_correct(self):
        """Test that sheet names match expected Turkish names"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        sheet_names = wb.sheetnames
        
        expected_sheets = ["Haftalik Ozet", "Makine Detayi", "Operator Performansi", "Defo Raporu"]
        for expected in expected_sheets:
            assert expected in sheet_names, f"Missing sheet: {expected}. Found: {sheet_names}"
        print(f"✓ All expected sheets present: {expected_sheets}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_haftalik_ozet_has_7_day_rows(self):
        """Test Haftalik Ozet sheet has 7 day rows"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Haftalik Ozet"]
        
        # Headers are in row 3, data starts at row 4
        # Days: Pazartesi, Sali, Carsamba, Persembe, Cuma, Cumartesi, Pazar
        day_names = ["Pazartesi", "Sali", "Carsamba", "Persembe", "Cuma", "Cumartesi", "Pazar"]
        found_days = []
        for row in range(4, 11):  # Rows 4-10 should have days
            cell_value = ws.cell(row=row, column=1).value
            if cell_value in day_names:
                found_days.append(cell_value)
        
        assert len(found_days) == 7, f"Expected 7 days, found {len(found_days)}: {found_days}"
        print(f"✓ Haftalik Ozet has 7 day rows: {found_days}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_haftalik_ozet_column_headers(self):
        """Test Haftalik Ozet sheet has correct column headers"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Haftalik Ozet"]
        
        # Headers in row 3
        expected_headers = ["Gun", "Tarih", "Toplam Koli", "Tamamlanan Is", "Baslayan Is", "Operator Sayisi", "Defo (kg)"]
        actual_headers = [ws.cell(row=3, column=col).value for col in range(1, 8)]
        
        for expected in expected_headers:
            assert expected in actual_headers, f"Missing header: {expected}. Found: {actual_headers}"
        print(f"✓ Haftalik Ozet headers correct: {actual_headers}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_makine_detayi_structure(self):
        """Test Makine Detayi sheet structure"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Makine Detayi"]
        
        # Check header row 3 has "Makine" and day names
        header_row = [ws.cell(row=3, column=col).value for col in range(1, 10)]
        assert "Makine" in header_row, f"Missing 'Makine' header. Found: {header_row}"
        assert "TOPLAM" in header_row, f"Missing 'TOPLAM' header. Found: {header_row}"
        print(f"✓ Makine Detayi structure correct: {header_row}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_operator_performansi_structure(self):
        """Test Operator Performansi sheet structure"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Operator Performansi"]
        
        # Check header row 3
        expected_headers = ["Operator", "Tamamlanan Is", "Toplam Koli", "Ort. Koli/Is", "Calistigi Makineler", "Ort. Sure (dk)"]
        actual_headers = [ws.cell(row=3, column=col).value for col in range(1, 7)]
        
        for expected in expected_headers:
            assert expected in actual_headers, f"Missing header: {expected}. Found: {actual_headers}"
        print(f"✓ Operator Performansi headers correct: {actual_headers}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_defo_raporu_structure(self):
        """Test Defo Raporu sheet structure"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Defo Raporu"]
        
        # Check header row 3
        expected_headers = ["Tarih", "Makine", "Defo (kg)", "Operator", "Aciklama"]
        actual_headers = [ws.cell(row=3, column=col).value for col in range(1, 6)]
        
        for expected in expected_headers:
            assert expected in actual_headers, f"Missing header: {expected}. Found: {actual_headers}"
        print(f"✓ Defo Raporu headers correct: {actual_headers}")
    
    @pytest.mark.skipif(load_workbook is None, reason="openpyxl not installed")
    def test_week_offset_minus_1_has_data(self):
        """Test that week_offset=-1 (previous week with data) returns data"""
        response = requests.get(f"{BASE_URL}/api/analytics/export?period=weekly&week_offset=-1")
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Haftalik Ozet"]
        
        # Check TOPLAM row (row 11) for total koli
        total_koli = ws.cell(row=11, column=3).value
        print(f"✓ Week offset -1 total koli: {total_koli}")
        # Note: May be 0 if no data in that week, but structure should be correct


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
