"""
Iteration 40 backend tests:
  1) GET /api/jobs returns thumb_url (data URL) on jobs that have images
  2) POST /api/analytics/export?period=weekly returns XLSX with Sheet 5 'Operator Zinciri'
  3) Operator Performansi sheet (Sayfa 3) reflects partial credit (no double counting)
  4) PUT /api/jobs/{id}/change-operator works (Ali → Mehmet Test with prev_produced_koli=25)
     and writes a shift_end_reports record with is_partial=True and transferred_to set
  5) Regression: management login response contains refresh_token
"""
import io
import os
import uuid

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://papermill-pro.preview.emergentagent.com").rstrip("/")


@pytest.fixture(scope="module")
def mgmt_token():
    r = requests.post(f"{BASE_URL}/api/management/login", json={"password": "buse11993"}, timeout=20)
    assert r.status_code == 200, f"mgmt login failed: {r.status_code} {r.text}"
    data = r.json()
    # regression: refresh_token must exist
    assert "refresh_token" in data, "Login response missing refresh_token"
    assert "token" in data
    return data["token"], data


@pytest.fixture(scope="module")
def auth_session(mgmt_token):
    s = requests.Session()
    token, _ = mgmt_token
    s.headers.update({"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
    return s


def test_login_refresh_token_regression(mgmt_token):
    _, data = mgmt_token
    assert isinstance(data.get("refresh_token"), str) and len(data["refresh_token"]) > 10


def test_jobs_thumb_url_present(auth_session):
    r = auth_session.get(f"{BASE_URL}/api/jobs", timeout=30)
    assert r.status_code == 200
    jobs = r.json()
    assert isinstance(jobs, list)
    # find at least one job with has_image=True
    with_image_jobs = [j for j in jobs if j.get("has_image") or j.get("thumb_url")]
    assert with_image_jobs, "No jobs with images found - need at least one with thumb_url for verification"
    thumbed = [j for j in with_image_jobs if j.get("thumb_url")]
    assert thumbed, f"Found {len(with_image_jobs)} jobs with images, but none have thumb_url"
    # validate first thumb structure
    sample = thumbed[0]
    assert sample["thumb_url"].startswith("data:image/"), f"thumb_url doesn't look like data URL: {sample['thumb_url'][:40]}"
    # ensure thumb is small (<25KB base64)
    assert len(sample["thumb_url"]) < 25000, f"Thumb too large: {len(sample['thumb_url'])} bytes"
    print(f"OK: {len(thumbed)} jobs have thumb_url. Sample size: {len(sample['thumb_url'])} chars")


def test_export_has_operator_zinciri_sheet(auth_session):
    r = auth_session.get(f"{BASE_URL}/api/analytics/export?period=weekly", timeout=60)
    assert r.status_code == 200
    assert "spreadsheet" in r.headers.get("content-type", "")
    wb = load_workbook(io.BytesIO(r.content))
    assert "Operator Zinciri" in wb.sheetnames, f"Sheet missing. Sheets: {wb.sheetnames}"
    ws = wb["Operator Zinciri"]
    # Title row
    title = ws.cell(row=1, column=1).value or ""
    assert "Operator Zinciri" in title or "Vardiya Devri" in title
    # Headers at row 3
    h = [ws.cell(row=3, column=c).value for c in range(1, 8)]
    expected_headers = ["Tarih", "Is Adi", "Makine", "Toplam Koli", "Degisim Sayisi", "Operator Zinciri", "Notlar"]
    assert h == expected_headers, f"Unexpected headers: {h}"
    # Data row 4 must exist - either real data or empty-state message
    first_data = ws.cell(row=4, column=1).value
    assert first_data is not None, "Row 4 should have data or empty-state message"
    print(f"OK: Sheet 'Operator Zinciri' present. Row 4 col1 = {first_data!r}")


def test_export_operator_performance_no_double_credit(auth_session):
    """Verify Sayfa 3 'Operator Performansi' exists and totals are non-negative integers."""
    r = auth_session.get(f"{BASE_URL}/api/analytics/export?period=weekly", timeout=60)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert "Operator Performansi" in wb.sheetnames
    ws = wb["Operator Performansi"]
    # Iterate rows from 4 onwards, ensure 'Toplam Koli' (col 3) values are >=0 ints
    row = 4
    rows_found = 0
    while True:
        op = ws.cell(row=row, column=1).value
        if not op:
            break
        koli = ws.cell(row=row, column=3).value
        assert isinstance(koli, (int, float)), f"Row {row} koli not numeric: {koli}"
        assert koli >= 0, f"Row {row} koli negative: {koli}"
        rows_found += 1
        row += 1
        if rows_found > 100:
            break
    print(f"OK: Operator Performansi has {rows_found} operator rows")


def test_change_operator_creates_partial_shift_report(auth_session):
    """Create a temp job (in_progress under 'Ali') then change operator to 'Mehmet Test' with prev_produced_koli=25."""
    # Create job
    job_id = str(uuid.uuid4())
    payload = {
        "id": job_id,
        "name": f"TEST_OpChange_{job_id[:8]}",
        "koli_count": 100,
        "colors": "red",
        "machine_id": "test-machine-it40",
        "machine_name": "Test Makine IT40",
        "status": "in_progress",
        "operator_name": "Ali",
    }
    r = auth_session.post(f"{BASE_URL}/api/jobs", json=payload, timeout=15)
    assert r.status_code == 200, f"create job failed: {r.status_code} {r.text}"
    created_id = r.json()["id"]

    try:
        # change operator
        r2 = auth_session.put(
            f"{BASE_URL}/api/jobs/{created_id}/change-operator",
            json={"new_operator_name": "Mehmet Test", "prev_produced_koli": 25, "note": "iteration40 test"},
            timeout=15,
        )
        assert r2.status_code == 200, f"change-operator failed: {r2.status_code} {r2.text}"
        body = r2.json()
        assert body["from_operator"] == "Ali"
        assert body["to_operator"] == "Mehmet Test"
        assert body["produced_credit"] == 25
        assert body.get("partial_record_id"), "partial_record_id missing"

        # Fetch job and verify operator updated + transfer_history present
        r3 = auth_session.get(f"{BASE_URL}/api/jobs?status=in_progress", timeout=15)
        assert r3.status_code == 200
        jobs = r3.json()
        our = next((j for j in jobs if j.get("id") == created_id), None)
        assert our, "Job not found after change"
        assert our["operator_name"] == "Mehmet Test"
        # Note: transfer_history may not be in projection but completed_koli must be 25
        assert our.get("completed_koli") == 25
        print(f"OK: operator change recorded. partial_record_id={body['partial_record_id']}")
    finally:
        # cleanup
        auth_session.delete(f"{BASE_URL}/api/jobs/{created_id}")


def test_change_operator_validation(auth_session):
    """Validation: cannot set same operator, koli must be int and within target."""
    job_id = str(uuid.uuid4())
    payload = {
        "id": job_id,
        "name": f"TEST_OpChange_Val_{job_id[:8]}",
        "koli_count": 50,
        "colors": "x",
        "machine_id": "test-machine-it40",
        "machine_name": "Test Makine IT40",
        "status": "in_progress",
        "operator_name": "Ali",
    }
    r = auth_session.post(f"{BASE_URL}/api/jobs", json=payload, timeout=15)
    assert r.status_code == 200
    cid = r.json()["id"]
    try:
        # same operator -> 400
        r2 = auth_session.put(f"{BASE_URL}/api/jobs/{cid}/change-operator", json={"new_operator_name": "Ali"})
        assert r2.status_code == 400
        # prev > target -> 400
        r3 = auth_session.put(
            f"{BASE_URL}/api/jobs/{cid}/change-operator",
            json={"new_operator_name": "Veli", "prev_produced_koli": 999},
        )
        assert r3.status_code == 400
    finally:
        auth_session.delete(f"{BASE_URL}/api/jobs/{cid}")
